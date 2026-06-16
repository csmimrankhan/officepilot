from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field, asdict
from datetime import date
from pathlib import Path
from typing import Any

logger = logging.getLogger("officepilot.accounting_report_comparison")

DEMO_CURRENT_MONTH = {
    "rows": [
        {"account": "Income", "amount": 250000.0, "type": "income"},
        {"account": "Cost of Goods Sold", "amount": 50000.0, "type": "expense"},
        {"account": "Salaries & Wages", "amount": 60000.0, "type": "expense"},
        {"account": "Rent & Utilities", "amount": 20000.0, "type": "expense"},
        {"account": "Office Supplies", "amount": 5000.0, "type": "expense"},
        {"account": "Marketing & Advertising", "amount": 10000.0, "type": "expense"},
        {"account": "Professional Fees", "amount": 5000.0, "type": "expense"},
    ],
    "total_income": 250000.0,
    "total_expenses": 150000.0,
    "net_income": 100000.0,
}

DEMO_PREVIOUS_MONTH = {
    "rows": [
        {"account": "Income", "amount": 220000.0, "type": "income"},
        {"account": "Cost of Goods Sold", "amount": 45000.0, "type": "expense"},
        {"account": "Salaries & Wages", "amount": 55000.0, "type": "expense"},
        {"account": "Rent & Utilities", "amount": 20000.0, "type": "expense"},
        {"account": "Office Supplies", "amount": 6000.0, "type": "expense"},
        {"account": "Marketing & Advertising", "amount": 8000.0, "type": "expense"},
        {"account": "Professional Fees", "amount": 6000.0, "type": "expense"},
    ],
    "total_income": 220000.0,
    "total_expenses": 140000.0,
    "net_income": 80000.0,
}


@dataclass
class PnLRow:
    account: str
    amount: float
    type: str  # "income" | "expense" | "gross_profit" | "net_income"


@dataclass
class PnLReport:
    rows: list[PnLRow] = field(default_factory=list)
    total_income: float = 0.0
    total_expenses: float = 0.0
    gross_profit: float | None = None
    net_income: float = 0.0
    period_label: str = ""


@dataclass
class LineDifference:
    account: str
    current_amount: float
    previous_amount: float
    difference: float
    percentage_change: float | None


@dataclass
class PnLComparison:
    current: PnLReport
    previous: PnLReport
    income_difference: float = 0.0
    expense_difference: float = 0.0
    net_income_difference: float = 0.0
    income_percentage_change: float | None = None
    expense_percentage_change: float | None = None
    net_income_percentage_change: float | None = None
    line_differences: list[LineDifference] = field(default_factory=list)


def get_demo_current_report() -> PnLReport:
    return PnLReport(
        rows=[PnLRow(**r) for r in DEMO_CURRENT_MONTH["rows"]],
        total_income=DEMO_CURRENT_MONTH["total_income"],
        total_expenses=DEMO_CURRENT_MONTH["total_expenses"],
        net_income=DEMO_CURRENT_MONTH["net_income"],
        gross_profit=DEMO_CURRENT_MONTH["total_income"] - sum(
            r["amount"] for r in DEMO_CURRENT_MONTH["rows"] if r["account"] == "Cost of Goods Sold"
        ),
        period_label="Current Month",
    )


def get_demo_previous_report() -> PnLReport:
    return PnLReport(
        rows=[PnLRow(**r) for r in DEMO_PREVIOUS_MONTH["rows"]],
        total_income=DEMO_PREVIOUS_MONTH["total_income"],
        total_expenses=DEMO_PREVIOUS_MONTH["total_expenses"],
        net_income=DEMO_PREVIOUS_MONTH["net_income"],
        gross_profit=DEMO_PREVIOUS_MONTH["total_income"] - sum(
            r["amount"] for r in DEMO_PREVIOUS_MONTH["rows"] if r["account"] == "Cost of Goods Sold"
        ),
        period_label="Previous Month",
    )


def detect_report_format(file_path: str) -> str:
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext == ".csv":
        return "csv"
    elif ext == ".xlsx":
        return "xlsx"
    elif ext == ".pdf":
        return "pdf"
    elif ext == ".json":
        return "json"
    elif ext == ".txt":
        return "txt"
    return "unknown"


def read_pnl_report(file_path: str) -> PnLReport:
    fmt = detect_report_format(file_path)
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if fmt == "json":
        return _read_pnl_json(path)
    elif fmt == "csv":
        return _read_pnl_csv(path)
    elif fmt == "xlsx":
        return _read_pnl_xlsx(path)
    elif fmt == "txt":
        return _read_pnl_txt(path)
    else:
        raise ValueError(f"Unsupported report format: {fmt}")


def _read_pnl_json(path: Path) -> PnLReport:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows_data = data.get("rows", data.get("data", []))
    rows = []
    for r in rows_data:
        rows.append(PnLRow(
            account=r.get("account", r.get("name", "")),
            amount=float(r.get("amount", 0)),
            type=r.get("type", _infer_type(r.get("account", ""))),
        ))
    return PnLReport(
        rows=rows,
        total_income=float(data.get("total_income", sum(r.amount for r in rows if r.type == "income"))),
        total_expenses=float(data.get("total_expenses", sum(r.amount for r in rows if r.type == "expense"))),
        net_income=float(data.get("net_income", data.get("net_profit", 0))),
        gross_profit=data.get("gross_profit"),
        period_label=data.get("period_label", data.get("period", "")),
    )


def _read_pnl_csv(path: Path) -> PnLReport:
    import csv
    rows = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            account = row.get("Account", row.get("account", row.get("Category", "")))
            amount = float(row.get("Amount", row.get("amount", 0)))
            type_val = row.get("Type", row.get("type", _infer_type(account)))
            rows.append(PnLRow(account=account, amount=amount, type=type_val))
    return _derive_report(rows)


def _read_pnl_xlsx(path: Path) -> PnLReport:
    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    rows = []
    header_row = None
    for row_cells in ws.iter_rows(min_row=1, values_only=True):
        if not any(c is not None for c in row_cells):
            continue
        if header_row is None:
            header_row = [str(c or "").lower().strip() for c in row_cells]
            continue
        row_data = dict(zip(header_row, [str(c or "") for c in row_cells]))
        account_col = next((k for k in row_data if k in ("account", "name", "category", "item")), None)
        amount_col = next((k for k in row_data if k in ("amount", "value", "total")), None)
        type_col = next((k for k in row_data if k in ("type", "classification")), None)
        if account_col is None:
            continue
        account = row_data[account_col]
        amount = float(row_data[amount_col]) if amount_col else 0.0
        type_val = row_data[type_col] if type_col else _infer_type(account)
        rows.append(PnLRow(account=account, amount=amount, type=type_val))
    return _derive_report(rows)


def _read_pnl_txt(path: Path) -> PnLReport:
    text = path.read_text(encoding="utf-8")
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("//"):
            continue
        parts = line.split()
        if len(parts) >= 2:
            amount_str = parts[-1]
            try:
                amount = float(amount_str.replace(",", ""))
                account = " ".join(parts[:-1])
                type_val = _infer_type(account)
                rows.append(PnLRow(account=account, amount=amount, type=type_val))
            except ValueError:
                continue
    return _derive_report(rows)


def _infer_type(account: str) -> str:
    account_lower = account.lower().strip()
    income_keywords = ("income", "revenue", "sales", "gross profit", "interest income", "other income")
    expense_keywords = ("expense", "cost", "salary", "wage", "rent", "utility", "supplies",
                       "marketing", "advertising", "professional fee", "depreciation",
                       "insurance", "travel", "meals", "entertainment", "maintenance",
                       "repairs", "tax", "commission", "shipping", "postage", "dues",
                       "subscription", "training", "education", "licence", "license",
                       "bank fee", "interest expense", "miscellaneous")
    net_keywords = ("net income", "net profit", "net loss", "total income", "total expense")
    for kw in net_keywords:
        if kw in account_lower:
            if "income" in kw or "profit" in kw:
                return "net_income"
            return "total"
    for kw in income_keywords:
        if kw in account_lower:
            return "income"
    for kw in expense_keywords:
        if kw in account_lower:
            return "expense"
    return "expense"


def _derive_report(rows: list[PnLRow]) -> PnLReport:
    income_items = [r for r in rows if r.type == "income"]
    expense_items = [r for r in rows if r.type == "expense"]
    total_income = sum(r.amount for r in income_items)
    total_expenses = sum(r.amount for r in expense_items)
    net_items = [r for r in rows if r.type == "net_income"]
    net_income = total_income - total_expenses
    if net_items:
        net_income = net_items[0].amount
    return PnLReport(
        rows=rows,
        total_income=total_income,
        total_expenses=total_expenses,
        net_income=net_income,
    )


def normalize_pnl_rows(rows: list[dict]) -> list[PnLRow]:
    return [PnLRow(**r) if isinstance(r, dict) else r for r in rows]


def _calc_pct_change(current: float, previous: float) -> float | None:
    if previous == 0:
        return None
    return ((current - previous) / abs(previous)) * 100


def compare_pnl_reports(current: PnLReport, previous: PnLReport) -> PnLComparison:
    line_diffs = []
    prev_by_account = {r.account: r for r in previous.rows}

    for cur_row in current.rows:
        prev_row = prev_by_account.pop(cur_row.account, None)
        prev_amount = prev_row.amount if prev_row else 0.0
        diff = cur_row.amount - prev_amount
        pct = _calc_pct_change(cur_row.amount, prev_amount)
        line_diffs.append(LineDifference(
            account=cur_row.account,
            current_amount=cur_row.amount,
            previous_amount=prev_amount,
            difference=diff,
            percentage_change=pct,
        ))

    for prev_row in prev_by_account.values():
        line_diffs.append(LineDifference(
            account=prev_row.account,
            current_amount=0.0,
            previous_amount=prev_row.amount,
            difference=-prev_row.amount,
            percentage_change=-100.0,
        ))

    income_diff = current.total_income - previous.total_income
    expense_diff = current.total_expenses - previous.total_expenses
    net_diff = current.net_income - previous.net_income

    return PnLComparison(
        current=current,
        previous=previous,
        income_difference=income_diff,
        expense_difference=expense_diff,
        net_income_difference=net_diff,
        income_percentage_change=_calc_pct_change(current.total_income, previous.total_income),
        expense_percentage_change=_calc_pct_change(current.total_expenses, previous.total_expenses),
        net_income_percentage_change=_calc_pct_change(current.net_income, previous.net_income),
        line_differences=line_diffs,
    )


def create_pnl_comparison_excel(
    current: PnLReport,
    previous: PnLReport,
    comparison: PnLComparison,
    output_path: str | Path,
) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Comparison"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_header(row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def write_row(row, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    write_header(1, ["Account", "Current Month", "Previous Month", "Difference", "Change %"])

    row = 2
    for ld in comparison.line_differences:
        pct_str = f"{ld.percentage_change:+.1f}%" if ld.percentage_change is not None else "N/A"
        write_row(row, [
            ld.account,
            round(ld.current_amount, 2),
            round(ld.previous_amount, 2),
            round(ld.difference, 2),
            pct_str,
        ])
        row += 1

    row += 1

    bold_font = Font(bold=True)
    write_row(row, ["Total Income",
                     round(current.total_income, 2),
                     round(previous.total_income, 2),
                     round(comparison.income_difference, 2),
                     f"{comparison.income_percentage_change:+.1f}%" if comparison.income_percentage_change is not None else "N/A"])
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = bold_font
    row += 1

    write_row(row, ["Total Expenses",
                     round(current.total_expenses, 2),
                     round(previous.total_expenses, 2),
                     round(comparison.expense_difference, 2),
                     f"{comparison.expense_percentage_change:+.1f}%" if comparison.expense_percentage_change is not None else "N/A"])
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = bold_font
    row += 1

    if current.gross_profit is not None or previous.gross_profit is not None:
        gp_diff = (current.gross_profit or 0) - (previous.gross_profit or 0)
        gp_pct = _calc_pct_change(current.gross_profit or 0, previous.gross_profit or 0)
        write_row(row, ["Gross Profit",
                         round(current.gross_profit or 0, 2),
                         round(previous.gross_profit or 0, 2),
                         round(gp_diff, 2),
                         f"{gp_pct:+.1f}%" if gp_pct is not None else "N/A"])
        for col in range(1, 6):
            ws.cell(row=row, column=col).font = bold_font
        row += 1

    net_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    write_row(row, ["Net Income",
                     round(current.net_income, 2),
                     round(previous.net_income, 2),
                     round(comparison.net_income_difference, 2),
                     f"{comparison.net_income_percentage_change:+.1f}%" if comparison.net_income_percentage_change is not None else "N/A"])
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.font = bold_font
        cell.fill = net_fill
    row += 2

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("P&L comparison Excel created: %s", output_path)
    return str(output_path)


def build_pnl_summary_text(comparison: PnLComparison, language: str = "en") -> str:
    ni_diff = comparison.net_income_difference
    ni_pct = comparison.net_income_percentage_change
    ni_pct_str = f"{ni_pct:+.1f}%" if ni_pct is not None else "N/A"
    sign = "increased" if ni_diff >= 0 else "decreased"
    abs_diff = abs(ni_diff)
    direction = "up" if ni_diff >= 0 else "down"

    current_period = comparison.current.period_label or "this month"
    previous_period = comparison.previous.period_label or "last month"

    income_pct = comparison.income_percentage_change
    expense_pct = comparison.expense_percentage_change
    income_pct_str = f"{income_pct:+.1f}%" if income_pct is not None else "N/A"
    expense_pct_str = f"{expense_pct:+.1f}%" if expense_pct is not None else "N/A"

    if language in ("ur", "roman_urdu"):
        return (
            f"Maine {current_period} ka P&L {previous_period} se compare kar diya hai. "
            f"Net income {abs_diff:,.2f} se {sign} hui hai, yani {ni_pct_str}. "
            f"Total income {income_pct_str}"
            f" aur total expenses {expense_pct_str}"
            f" hain."
        )

    pct_line = f", which is {ni_pct_str}" if ni_pct is not None else ""
    return (
        f"I compared {current_period}'s P&L with {previous_period}. "
        f"Net income {sign} by {abs_diff:,.2f}{pct_line}. "
        f"Total income changed by {income_pct_str} "
        f"and total expenses changed by {expense_pct_str}."
    )


def pnl_comparison_to_dict(comparison: PnLComparison) -> dict[str, Any]:
    return {
        "current": {
            "total_income": comparison.current.total_income,
            "total_expenses": comparison.current.total_expenses,
            "gross_profit": comparison.current.gross_profit,
            "net_income": comparison.current.net_income,
            "period_label": comparison.current.period_label,
            "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in comparison.current.rows],
        },
        "previous": {
            "total_income": comparison.previous.total_income,
            "total_expenses": comparison.previous.total_expenses,
            "gross_profit": comparison.previous.gross_profit,
            "net_income": comparison.previous.net_income,
            "period_label": comparison.previous.period_label,
            "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in comparison.previous.rows],
        },
        "comparison": {
            "income_difference": comparison.income_difference,
            "expense_difference": comparison.expense_difference,
            "net_income_difference": comparison.net_income_difference,
            "income_percentage_change": comparison.income_percentage_change,
            "expense_percentage_change": comparison.expense_percentage_change,
            "net_income_percentage_change": comparison.net_income_percentage_change,
        },
        "line_differences": [
            {
                "account": ld.account,
                "current_amount": ld.current_amount,
                "previous_amount": ld.previous_amount,
                "difference": ld.difference,
                "percentage_change": ld.percentage_change,
            }
            for ld in comparison.line_differences
        ],
    }
