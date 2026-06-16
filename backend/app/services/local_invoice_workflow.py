from __future__ import annotations

import json
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger("officepilot.local_invoice_workflow")

INVOICE_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".csv", ".xlsx", ".xls", ".txt"}
INVOICE_KEYWORDS = re.compile(r"(invoice|bill|receipt|payment)", re.IGNORECASE)


class InvoiceFileInfo:
    def __init__(self, path: str, filename: str, modified: str, size: int):
        self.path = path
        self.filename = filename
        self.modified = modified
        self.size = size


class ExtractedInvoice:
    def __init__(
        self,
        vendor: str = "",
        invoice_number: str = "",
        invoice_date: str = "",
        total_amount: float = 0.0,
        tax: float = 0.0,
        currency: str = "USD",
        source_file: str = "",
        confidence: float = 0.0,
        warnings: list[str] | None = None,
        status: str = "imported",
    ):
        self.vendor = vendor
        self.invoice_number = invoice_number
        self.invoice_date = invoice_date
        self.total_amount = total_amount
        self.tax = tax
        self.currency = currency
        self.source_file = source_file
        self.confidence = confidence
        self.warnings = warnings or []
        self.status = status


def scan_folder_for_invoices(
    folder_path: str | Path,
    date_filter: str = "today",
    keywords: bool = True,
) -> list[InvoiceFileInfo]:
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        logger.warning("Folder not found: %s", folder_path)
        return []

    today = date.today()
    results: list[InvoiceFileInfo] = []

    for entry in folder.iterdir():
        if not entry.is_file():
            continue
        ext = entry.suffix.lower()
        if ext not in INVOICE_EXTENSIONS:
            continue
        if keywords and not INVOICE_KEYWORDS.search(entry.stem):
            continue

        mtime = datetime.fromtimestamp(entry.stat().st_mtime).date()

        if date_filter == "today" and mtime != today:
            continue
        if date_filter == "yesterday" and mtime != today.replace(day=today.day - 1):
            continue

        results.append(InvoiceFileInfo(
            path=str(entry.resolve()),
            filename=entry.name,
            modified=mtime.isoformat(),
            size=entry.stat().st_size,
        ))

    results.sort(key=lambda f: f.filename)
    logger.info("Scanned %s: found %d invoice files", folder_path, len(results))
    return results


def extract_invoice_from_file(file_path: str) -> ExtractedInvoice:
    path = Path(file_path)
    if not path.exists():
        return ExtractedInvoice(
            source_file=file_path,
            confidence=0.0,
            warnings=["File not found"],
            status="needs_review",
        )

    text = ""
    ext = path.suffix.lower()

    try:
        if ext == ".pdf":
            text = _extract_pdf_text(path)
        elif ext in (".png", ".jpg", ".jpeg"):
            text = _extract_image_text(path)
        elif ext in (".csv", ".xlsx", ".xls"):
            return _extract_spreadsheet(path)
        elif ext == ".txt":
            text = path.read_text(encoding="utf-8", errors="replace")
        else:
            text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        logger.error("Extraction failed for %s: %s", file_path, e)
        return ExtractedInvoice(
            source_file=file_path,
            confidence=0.0,
            warnings=[f"Extraction error: {e}"],
            status="needs_review",
        )

    if not text.strip():
        return ExtractedInvoice(
            source_file=file_path,
            confidence=0.0,
            warnings=["No extractable text found"],
            status="needs_review",
        )

    return _parse_invoice_text(text, file_path)


def _extract_pdf_text(path: Path) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            pages = [p.extract_text() for p in pdf.pages if p.extract_text()]
            return "\n".join(pages)
    except ImportError:
        try:
            import PyPDF2
            text = []
            with open(str(path), "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        text.append(t)
            return "\n".join(text)
        except ImportError:
            try:
                import pdfminer
                from pdfminer.high_level import extract_text
                return extract_text(str(path))
            except ImportError:
                return path.read_text(encoding="utf-8", errors="replace")


def _extract_image_text(path: Path) -> str:
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(str(path))
        return pytesseract.image_to_string(img)
    except ImportError:
        return path.read_text(encoding="utf-8", errors="replace")


def _extract_spreadsheet(path: Path) -> ExtractedInvoice:
    import openpyxl
    wb = None
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(c or "") for c in row]
            if any(v.strip() for v in vals):
                rows.append(vals)

        if not rows:
            return ExtractedInvoice(
                source_file=str(path),
                confidence=0.0,
                warnings=["Empty spreadsheet"],
                status="needs_review",
            )

        headers = [h.lower().strip() for h in rows[0]]
        data_rows = rows[1:] if len(rows) > 1 else []

        total = 0.0
        vendor = path.stem
        invoice_no = ""
        tax = 0.0
        currency = "USD"

        for col_idx, h in enumerate(headers):
            col_vals = [r[col_idx] if col_idx < len(r) else "" for r in data_rows]
            num_vals = []
            for v in col_vals:
                try:
                    num_vals.append(float(v.replace(",", "").replace("$", "").replace(" ", "")))
                except (ValueError, AttributeError):
                    pass
            if "total" in h or "amount" in h:
                total = sum(num_vals) if num_vals else 0.0
            elif "tax" in h or "vat" in h or "gst" in h:
                tax = sum(num_vals) if num_vals else 0.0
            elif "vendor" in h or "name" in h or "customer" in h:
                if col_vals and col_vals[0].strip():
                    vendor = col_vals[0].strip()
            elif "invoice" in h and ("no" in h or "number" in h or "#" in h):
                if col_vals and col_vals[0].strip():
                    invoice_no = col_vals[0].strip()
            elif "currency" in h:
                if col_vals and col_vals[0].strip():
                    currency = col_vals[0].strip().upper()

        confidence = 0.5 if total > 0 else 0.3
        warnings = []
        if total == 0:
            warnings.append("total_amount could not be determined")
        if not invoice_no:
            warnings.append("invoice_number missing")

        return ExtractedInvoice(
            vendor=vendor,
            invoice_number=invoice_no,
            total_amount=total,
            tax=tax,
            currency=currency,
            source_file=str(path),
            confidence=confidence,
            warnings=warnings,
            status="needs_review" if total == 0 else "ready_for_approval",
        )
    except ImportError:
        return ExtractedInvoice(
            source_file=str(path),
            confidence=0.0,
            warnings=["openpyxl not available to read spreadsheet"],
            status="needs_review",
        )
    except Exception as e:
        return ExtractedInvoice(
            source_file=str(path),
            confidence=0.0,
            warnings=[f"Spreadsheet error: {e}"],
            status="needs_review",
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _parse_invoice_text(text: str, file_path: str) -> ExtractedInvoice:
    from .parser import parse_invoice_text as _parse

    parsed = _parse(text)

    status = "imported"
    if parsed.confidence < 0.4:
        status = "needs_review"
    elif parsed.confidence < 0.7:
        status = "imported"
    else:
        status = "ready_for_approval"

    return ExtractedInvoice(
        vendor=parsed.vendor_name or "",
        invoice_number=parsed.invoice_number or "",
        invoice_date=parsed.invoice_date or "",
        total_amount=parsed.total_amount or 0.0,
        tax=parsed.tax or 0.0,
        currency=parsed.currency or "USD",
        source_file=file_path,
        confidence=parsed.confidence,
        warnings=parsed.warnings,
        status=status,
    )


def create_daily_invoices_excel(
    invoices: list[ExtractedInvoice],
    output_dir: str | Path,
) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    today_str = date.today().strftime("%Y_%m_%d")
    filename = f"Daily_Invoices_{today_str}.xlsx"
    output_path = Path(output_dir) / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Invoice Detail sheet
    ws = wb.active
    ws.title = "Invoices"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    warning_font = Font(color="CC6600")

    headers = [
        "Vendor", "Invoice Number", "Invoice Date", "Tax",
        "Total", "Currency", "Source File", "Status", "Warning"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    success_count = 0
    failed_count = 0
    review_count = 0
    total_amount = 0.0
    total_tax = 0.0

    for row_idx, inv in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=inv.vendor).border = thin_border
        ws.cell(row=row_idx, column=2, value=inv.invoice_number).border = thin_border
        ws.cell(row=row_idx, column=3, value=inv.invoice_date).border = thin_border
        ws.cell(row=row_idx, column=4, value=inv.tax if inv.tax else 0).border = thin_border
        ws.cell(row=row_idx, column=5, value=inv.total_amount).border = thin_border
        ws.cell(row=row_idx, column=6, value=inv.currency).border = thin_border
        ws.cell(row=row_idx, column=7, value=inv.source_file).border = thin_border
        ws.cell(row=row_idx, column=8, value=inv.status).border = thin_border
        warning_cell = ws.cell(row=row_idx, column=9, value="; ".join(inv.warnings) if inv.warnings else "")
        warning_cell.border = thin_border
        if inv.warnings:
            warning_cell.font = warning_font

        if inv.confidence >= 0.4 and inv.total_amount > 0:
            success_count += 1
        elif inv.confidence < 0.4:
            failed_count += 1
        if inv.status == "needs_review":
            review_count += 1
        total_amount += inv.total_amount
        total_tax += inv.tax or 0

    for col_idx in range(1, 10):
        ws.cell(row=len(invoices) + 2, column=col_idx).border = thin_border

    col_widths = [25, 18, 16, 12, 14, 10, 50, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    bold_font = Font(bold=True)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("Metric", "Value"),
        ("Invoice Count", len(invoices)),
        ("Successfully Extracted", success_count),
        ("Failed Extraction", failed_count),
        ("Needs Review", review_count),
        ("Total Amount", round(total_amount, 2)),
        ("Total Tax", round(total_tax, 2)),
        ("", ""),
        ("Date", today_str),
        ("Source Files", str(output_path.parent)),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        elif label == "Total Amount":
            c1.font = bold_font
            c2.font = bold_font

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    wb.save(str(output_path))
    logger.info("Daily invoices Excel created: %s with %d invoices", output_path, len(invoices))
    return str(output_path)


def build_folder_invoice_summary_text(
    count: int,
    success_count: int,
    total: float,
    excel_path: str,
    language: str = "en",
) -> str:
    total_str = f"{total:,.2f}"
    if language in ("ur", "roman_urdu"):
        return (
            f"Maine aaj ki {count} invoice files find ki hain. "
            f"{success_count} extract ho gayi hain. "
            f"Total {total_str} hai. "
            f"Excel file yahan save ho gayi hai: {excel_path}."
        )

    return (
        f"I found {count} invoice files today. "
        f"I extracted {success_count} successfully. "
        f"Total is {total_str}. "
        f"I saved the Excel file here: {excel_path}."
    )
