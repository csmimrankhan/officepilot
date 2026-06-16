from __future__ import annotations

import logging
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("officepilot.excel_tools")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None


VALID_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".csv"}

ACCOUNTING_COLUMN_PATTERNS = {
    "date": [r"date", r"dated", r"transaction\s*date", r"post\s*date", r"entry\s*date", r"invoice\s*date"],
    "vendor": [r"vendor", r"name", r"customer", r"client", r"supplier", r"payee", r"party", r"creditor", r"debitor"],
    "category": [r"category", r"type", r"account", r"class", r"department", r"group", r"classification", r"head"],
    "description": [r"description", r"memo", r"note", r"particulars", r"details", r"narrative", r"reference"],
    "amount": [r"amount", r"total", r"sum", r"value", r"amt", r"amount\s*\(.*\)", r"amount\s*in"],
    "debit": [r"debit", r"dr\b", r"debit\s*amount"],
    "credit": [r"credit", r"cr\b", r"credit\s*amount"],
}


def _match_accounting_column(header: str) -> str | None:
    hl = header.strip().lower()
    for col_type, patterns in ACCOUNTING_COLUMN_PATTERNS.items():
        for pat in patterns:
            import re
            if re.search(pat, hl):
                return col_type
    return None


_WORKSPACE_DIRS: list[str] = []


def set_workspace_dirs(dirs: list[str]):
    global _WORKSPACE_DIRS
    _WORKSPACE_DIRS = list(dirs)


def _is_path_allowed(path: str) -> bool:
    if not _WORKSPACE_DIRS:
        return True
    resolved = os.path.abspath(os.path.normpath(path))
    for d in _WORKSPACE_DIRS:
        if resolved.startswith(os.path.abspath(os.path.normpath(d))):
            return True
    return False


def _backup_path(original_path: str) -> str:
    p = Path(original_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(p.parent / f"{p.stem}_backup_{ts}{p.suffix}")


def _snapshot_path(file_path: str, data_dir: str | None = None) -> Path:
    from datetime import date
    if data_dir is None:
        data_dir = os.environ.get("OFFICEPILOT_DATA_DIR", os.path.join(os.path.dirname(__file__), "..", "..", "data"))
    today = date.today()
    snap_dir = Path(data_dir) / "snapshots" / "excel" / str(today.year) / f"{today.month:02d}" / f"{today.day:02d}"
    snap_dir.mkdir(parents=True, exist_ok=True)
    src = Path(file_path)
    import uuid
    dest = snap_dir / f"{uuid.uuid4().hex}{src.suffix}"
    shutil.copy2(src, dest)
    return dest


def create_backup_and_snapshot(file_path: str, data_dir: str | None = None) -> dict:
    if not os.path.isfile(file_path):
        return {"backup_created": False, "snapshot_created": False, "error": "File not found"}
    backup = _backup_path(file_path)
    shutil.copy2(file_path, backup)
    snap = _snapshot_path(file_path, data_dir)
    return {"backup_created": True, "backup_path": backup, "snapshot_created": True, "snapshot_path": str(snap)}


# ── Workbook Operations ──────────────────────────────────────────────

def create_workbook(path: str, sheets: list[str] | None = None) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in (sheets or ["Sheet1"]):
        wb.create_sheet(title=name)
    wb.save(path)
    wb.close()
    return {"success": True, "path": os.path.abspath(path), "sheets": sheets or ["Sheet1"]}


def open_workbook(path: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not os.path.isfile(path):
        return {"success": False, "error": "File not found"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        info = {
            "path": os.path.abspath(path),
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }
        wb.close()
        return {"success": True, **info}
    except Exception as e:
        return {"success": False, "error": str(e)}


def read_sheet(path: str, sheet_name: str, max_rows: int = 0) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not os.path.isfile(path):
        return {"success": False, "error": "File not found"}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows and i >= max_rows:
                break
            rows.append(list(row))
        wb.close()
        return {"success": True, "sheet_name": sheet_name, "rows": rows, "row_count": len(rows)}
    except Exception as e:
        return {"success": False, "error": str(e)}


def detect_columns(path: str, sheet_name: str) -> dict:
    result = read_sheet(path, sheet_name, max_rows=100)
    if not result.get("success"):
        return result
    rows = result.get("rows", [])
    if not rows:
        return {"success": True, "sheet_name": sheet_name, "columns": [], "row_count": 0}
    headers = rows[0] if rows else []
    data_rows = rows[1:]
    detected = []
    for i, h in enumerate(headers):
        col_letter = get_column_letter(i + 1)
        sample_values = [str(r[i]) for r in data_rows[:5] if i < len(r) and r[i] is not None]
        col_type = _infer_column_type(sample_values)
        detected.append({
            "index": i,
            "letter": col_letter,
            "header": str(h) if h is not None else f"Column {col_letter}",
            "sample_values": sample_values[:3],
            "inferred_type": col_type,
            "non_empty_count": sum(1 for r in data_rows if i < len(r) and r[i] is not None),
        })
    return {
        "success": True,
        "sheet_name": sheet_name,
        "columns": detected,
        "total_rows": len(data_rows),
        "header_row": headers,
    }


def _infer_column_type(samples: list[str]) -> str:
    import re
    if not samples:
        return "unknown"
    digit_count = sum(1 for s in samples if re.match(r'^[\d,.]+\s*$', s.strip()) if s)
    if digit_count > len(samples) // 2:
        return "number"
    date_count = 0
    for s in samples:
        if not s:
            continue
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                datetime.strptime(s.strip(), fmt)
                date_count += 1
                break
            except ValueError:
                continue
    if date_count > len(samples) // 2:
        return "date"
    return "text"


def write_cell(path: str, sheet_name: str, cell: str, value: Any) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(path)
        wb.close()
        return {"success": True, "cell": cell, "value": value, **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


def write_range(path: str, sheet_name: str, start_cell: str, values: list[list[Any]]) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                cell_ref = f"{get_column_letter(c_idx + 1)}{int(openpyxl.utils.cell.coordinate_from_string(start_cell)[1]) + r_idx}"
                ws[cell_ref] = val
        wb.save(path)
        wb.close()
        return {"success": True, "start_cell": start_cell, "rows_written": len(values), **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


def append_rows(path: str, sheet_name: str, rows: list[list[Any]]) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        for row in rows:
            ws.append(row)
        wb.save(path)
        wb.close()
        return {"success": True, "rows_appended": len(rows), **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


def save_workbook(path: str, save_as: str | None = None) -> dict:
    if not os.path.isfile(path):
        return {"success": False, "error": "Source file not found"}
    dest = save_as or path
    if dest != path:
        backup = create_backup_and_snapshot(dest) if os.path.isfile(dest) else {"backup_created": False, "snapshot_created": False}
        shutil.copy2(path, dest)
        return {"success": True, "path": os.path.abspath(dest), **backup}
    return {"success": True, "path": os.path.abspath(path), "saved": True}


# ── Formula Operations ───────────────────────────────────────────────

def apply_formula(path: str, sheet_name: str, cell: str, formula: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    from .excel_formula_compat import validate_formula_safety
    safe, reason = validate_formula_safety(formula)
    if not safe:
        return {"success": False, "error": reason}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        if not formula.startswith("="):
            formula = "=" + formula
        ws[cell] = formula
        wb.save(path)
        wb.close()
        return {"success": True, "cell": cell, "formula": formula, **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


def apply_total_formula(path: str, sheet_name: str, column_letter: str, start_row: int = 2, end_row: int = None) -> dict:
    from .excel_formula_compat import choose_formula
    if end_row is None:
        data = read_sheet(path, sheet_name)
        if not data.get("success"):
            return data
        end_row = data.get("row_count", 100)
    formula_template = choose_formula("sum")
    formula = formula_template.replace("{range}", f"{column_letter}{start_row}:{column_letter}{end_row}")
    cell = f"{column_letter}{end_row + 1}"
    return apply_formula(path, sheet_name, cell, formula)


def apply_sumif_formula(path: str, sheet_name: str, result_cell: str, criteria_range: str, criteria: str, sum_range: str) -> dict:
    from .excel_formula_compat import choose_formula
    template = choose_formula("sumif")
    formula = template.replace("{criteria_range}", criteria_range).replace("{criteria}", criteria).replace("{sum_range}", sum_range)
    return apply_formula(path, sheet_name, result_cell, formula)


def apply_lookup_formula(path: str, sheet_name: str, result_cell: str, lookup_value: str, table_array: str, return_col: str, col_index: int = 2, compatibility_mode: str = "excel_2016") -> dict:
    from .excel_formula_compat import choose_formula, ExcelCompatibilityMode
    mode = ExcelCompatibilityMode(compatibility_mode)
    template = choose_formula("find", mode)
    if "XLOOKUP" in template:
        formula = template.replace("{lookup_value}", lookup_value).replace("{lookup_array}", f"{table_array}:{table_array}").replace("{return_array}", return_col)
    elif "INDEX" in template:
        formula = template.replace("{return_column}", return_col).replace("{lookup_value}", lookup_value).replace("{lookup_column}", table_array)
    else:
        formula = template.replace("{lookup_value}", lookup_value).replace("{table_array}", table_array).replace("{col_index_num}", str(col_index))
    return apply_formula(path, sheet_name, result_cell, formula)


def verify_formula(path: str, sheet_name: str, cell: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        formula = ws[cell].value
        from .excel_formula_compat import validate_formula_safety
        safe, reason = validate_formula_safety(str(formula))
        wb.close()
        if not safe:
            return {"success": False, "formula": str(formula), "error": reason}
        return {"success": True, "cell": cell, "formula": str(formula), "valid": formula is not None}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── Formatting Operations ────────────────────────────────────────────

def _style_header_fill():
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _style_header_font():
    return Font(bold=True, color="FFFFFF", size=11)


def _thin_border():
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def format_header(path: str, sheet_name: str, header_row: int = 1) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        for cell in ws[header_row]:
            cell.font = _style_header_font()
            cell.fill = _style_header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name, "header_row": header_row}
    except Exception as e:
        return {"success": False, "error": str(e)}


def auto_size_columns(path: str, sheet_name: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_length:
                        max_length = len(val)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name}
    except Exception as e:
        return {"success": False, "error": str(e)}


def freeze_top_row(path: str, sheet_name: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name, "freeze_panes": "A2"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def apply_currency_format(path: str, sheet_name: str, column_letter: str, rows_range: tuple[int, int] = None) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        start, end = rows_range or (1, ws.max_row or 100)
        for row in range(start, end + 1):
            cell = ws[f"{column_letter}{row}"]
            cell.number_format = '#,##0.00'
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name, "column": column_letter}
    except Exception as e:
        return {"success": False, "error": str(e)}


def apply_date_format(path: str, sheet_name: str, column_letter: str, date_format: str = "YYYY-MM-DD") -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(column_letter), max_col=openpyxl.utils.column_index_from_string(column_letter)):
            for cell in row:
                cell.number_format = date_format
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name, "column": column_letter, "format": date_format}
    except Exception as e:
        return {"success": False, "error": str(e)}


def highlight_negative_values(path: str, sheet_name: str, column_letter: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        red_font = Font(color="FF0000", bold=True)
        for row in ws.iter_rows(min_col=openpyxl.utils.column_index_from_string(column_letter), max_col=openpyxl.utils.column_index_from_string(column_letter)):
            for cell in row:
                try:
                    if cell.value is not None and float(cell.value) < 0:
                        cell.font = red_font
                except (ValueError, TypeError):
                    pass
        wb.save(path)
        wb.close()
        return {"success": True, "sheet": sheet_name, "column": column_letter}
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_total_row(path: str, sheet_name: str, amount_column: str, label: str = "Total") -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        ws = wb[sheet_name]
        last_row = ws.max_row or 1
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value=label)
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        col_idx = openpyxl.utils.column_index_from_string(amount_column)
        total_cell = ws.cell(row=total_row, column=col_idx)
        total_cell.value = f"=SUM({amount_column}2:{amount_column}{last_row})"
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        wb.save(path)
        wb.close()
        return {"success": True, "total_row": total_row, "column": amount_column, **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── File Validation ─────────────────────────────────────────────────

def validate_file(path: str) -> dict:
    """Validate that the file exists and has a supported extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in VALID_EXTENSIONS:
        return {"valid": False, "error": f"Unsupported file type '{ext}'. Supported: {', '.join(sorted(VALID_EXTENSIONS))}"}
    if not os.path.isfile(path):
        return {"valid": False, "error": f"File not found: {path}"}
    return {"valid": True, "path": os.path.abspath(path), "extension": ext}


# ── Enhanced Accounting Column Detection ─────────────────────────────────

def detect_columns_with_semantics(path: str, sheet_name: str) -> dict:
    """Detect columns and map them to accounting semantic types."""
    result = detect_columns(path, sheet_name)
    if not result.get("success"):
        return result
    columns = result.get("columns", [])
    for col in columns:
        header = col.get("header", "")
        semantic = _match_accounting_column(header)
        col["semantic_type"] = semantic
        col["detected"] = semantic is not None
    return result


def suggest_summary_columns(path: str, sheet_name: str) -> dict:
    """Auto-detect which column to group by and which to sum."""
    result = detect_columns_with_semantics(path, sheet_name)
    if not result.get("success"):
        return result
    columns = result.get("columns", [])
    group_candidates = []
    vendor_candidates = []
    date_candidates = []
    value_candidates = []
    for col in columns:
        st = col.get("semantic_type")
        inferred = col.get("inferred_type", "text")
        if st == "category":
            group_candidates.append(col)
        elif st in ("vendor", "description"):
            vendor_candidates.append(col)
        elif st == "date":
            date_candidates.append(col)
        if st in ("amount", "debit", "credit", "total"):
            value_candidates.append(col)
        if st is None and inferred == "number":
            value_candidates.append(col)

    group_col = group_candidates[0] if group_candidates else (vendor_candidates[0] if vendor_candidates else (date_candidates[0] if date_candidates else (columns[0] if columns else None)))
    value_col = value_candidates[0] if value_candidates else (columns[-1] if len(columns) > 1 else (columns[0] if columns else None))

    return {
        "success": True,
        "sheet_name": sheet_name,
        "columns": columns,
        "suggested_group_by": group_col["header"] if group_col else None,
        "suggested_group_by_letter": group_col["letter"] if group_col else None,
        "suggested_value": value_col["header"] if value_col else None,
        "suggested_value_letter": value_col["letter"] if value_col else None,
        "total_rows": result.get("total_rows", 0),
    }


# ── High-Level Summary Orchestration ─────────────────────────────────────

def create_summary_from_file(path: str, options: dict | None = None) -> dict:
    """High-level function: validate, detect, summarize, format, save copy.

    Runs the full pipeline:
      1. validate_file
      2. create_backup_and_snapshot
      3. suggest_summary_columns (auto-detect group/value)
      4. create_summary_sheet (or auto-group by detected column)
      5. format the summary sheet
      6. save a copy (never overwrites original unless explicit)
    """
    opts = options or {}
    mode = opts.get("mode", "live")
    source_sheet = opts.get("source_sheet") or opts.get("sheet_name", "")
    group_by = opts.get("group_by_column") or ""
    value_col = opts.get("value_column") or opts.get("amount_column", "")

    # 1. Validate
    val = validate_file(path)
    if not val["valid"]:
        return {"status": "failed", "error": val["error"]}

    # Auto-detect first sheet if source_sheet is empty
    if not source_sheet:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            if wb.sheetnames:
                source_sheet = wb.sheetnames[0]
            wb.close()
        except Exception:
            pass

    if mode == "dry_run":
        return _dry_run_summary(path, source_sheet, group_by, value_col, opts)

    # 2. Auto-detect columns if not specified
    if not group_by or not value_col:
        suggestion = suggest_summary_columns(path, source_sheet or "")
        if not suggestion.get("success"):
            return {"status": "failed", "error": suggestion.get("error", "Could not read file")}
        if not group_by and suggestion.get("suggested_group_by"):
            group_by = suggestion["suggested_group_by"]
        if not value_col and suggestion.get("suggested_value"):
            value_col = suggestion["suggested_value"]
        if not source_sheet:
            source_sheet = suggestion.get("sheet_name", "Sheet1")
        if not group_by or not value_col:
            return {
                "status": "needs_input",
                "input_type": "column_picker",
                "columns": suggestion.get("columns", []),
                "message": "Could not auto-detect columns. Which column should I group by and which should I sum?",
            }

    # 3. Backup original (before any modification)
    backup = create_backup_and_snapshot(path)
    if not backup.get("backup_created"):
        return {"status": "failed", "error": backup.get("error", "Backup failed")}

    # 4. Save output copy FIRST (copy of original, never modify original)
    output = save_output_copy(path, source_sheet)
    if not output.get("success"):
        return {"status": "failed", "error": output.get("error", "Save copy failed")}
    output_path = output["output_path"]

    # 5. Read raw data from original for total computation
    read_result = read_sheet(path, source_sheet)
    data_rows = read_result.get("rows", []) if read_result.get("success") else []
    headers = data_rows[0] if data_rows else []
    raw_data = data_rows[1:] if len(data_rows) > 1 else []

    # 6. Create summary sheet on the OUTPUT COPY (not the original)
    summary = create_summary_sheet(output_path, source_sheet, group_by, value_col)
    if not summary.get("success"):
        return {"status": "failed", "error": summary.get("error", "Summary creation failed")}

    # 7. Format summary sheet on output copy
    summary_sheet_name = summary.get("summary_sheet", f"{source_sheet[:20]}_Summary")
    format_header(output_path, summary_sheet_name)
    auto_size_columns(output_path, summary_sheet_name)
    freeze_top_row(output_path, summary_sheet_name)

    # 8. Apply currency format to value column
    if "suggested_value_letter" in summary:
        col_letter = summary["suggested_value_letter"]
    else:
        for c in (suggest_summary_columns(output_path, source_sheet).get("columns", []) if not group_by else []):
            if c.get("header") == value_col:
                col_letter = c["letter"]
                break
        else:
            col_letter = "C"
    try:
        apply_currency_format(output_path, summary_sheet_name, col_letter, (2, summary.get("group_count", 10) + 2))
    except Exception:
        pass

    # 9. Add grand total row on summary
    try:
        total_col_letter = col_letter
        add_total_row(output_path, summary_sheet_name, total_col_letter, "Grand Total")
    except Exception:
        pass

    # 10. Compute total from original data
    total = 0.0
    val_idx = None
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == value_col.strip().lower():
            val_idx = i
            break
    if val_idx is not None:
        for row in raw_data:
            if val_idx < len(row):
                try:
                    total += float(row[val_idx]) if row[val_idx] is not None else 0.0
                except (ValueError, TypeError):
                    pass

    return {
        "status": "success",
        "input_path": os.path.abspath(path),
        "backup_path": backup.get("backup_path", ""),
        "output_path": output_path,
        "summary_sheet": summary_sheet_name,
        "source_sheet": source_sheet,
        "detected_group_by": group_by,
        "detected_value": value_col,
        "rows_processed": len(raw_data),
        "total": round(total, 2),
        "warnings": [],
    }


def _dry_run_summary(path: str, source_sheet: str, group_by: str, value_col: str, opts: dict) -> dict:
    """Return a preview of what would happen, without modifying any files."""
    val = validate_file(path)
    if not val["valid"]:
        return {"mode": "dry_run", "would_modify_file": False, "error": val["error"]}

    ss = source_sheet or ""
    if not ss:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            if wb.sheetnames:
                ss = wb.sheetnames[0]
            wb.close()
        except Exception:
            pass

    suggestion = suggest_summary_columns(path, ss)
    if not suggestion.get("success"):
        return {"mode": "dry_run", "would_modify_file": False, "error": suggestion.get("error", "Could not read file")}

    gb = group_by or suggestion.get("suggested_group_by") or ""
    vc = value_col or suggestion.get("suggested_value") or ""
    ss = source_sheet or suggestion.get("sheet_name", "Sheet1")

    columns = suggestion.get("columns", [])
    total_rows = suggestion.get("total_rows", 0)

    # Preview first few data rows
    read_result = read_sheet(path, ss, max_rows=6)
    preview_rows = []
    all_rows = read_result.get("rows", [])
    if all_rows:
        preview_rows = all_rows[:6]

    planned_actions = [
        {"action": "Create backup", "detail": f"Backup of {os.path.basename(path)}"},
        {"action": f"Group by '{gb}'", "detail": f"Summarize rows grouped by column: {gb}"},
        {"action": f"Sum '{vc}'", "detail": f"Calculate total of column: {vc}"},
        {"action": "Create Summary sheet", "detail": f"New sheet '{ss}_Summary' with grouped totals"},
        {"action": "Add grand total row", "detail": "Add Grand Total row with SUM formula"},
        {"action": "Format report", "detail": "Bold headers, freeze top row, auto-size columns, currency format"},
        {"action": "Save output copy", "detail": f"Save as {os.path.basename(path).replace('.xlsx', '')}_summary_*.xlsx"},
    ]

    return {
        "mode": "dry_run",
        "would_modify_file": False,
        "planned_actions": planned_actions,
        "detected_columns": columns,
        "preview_rows": preview_rows,
        "total_data_rows": total_rows,
        "suggested_group_by": gb,
        "suggested_value": vc,
        "source_sheet": ss,
    }


def save_output_copy(path: str, source_sheet: str = "Source") -> dict:
    """Save a copy of the workbook as a new output file.

    Never overwrites original. Pattern: originalname_summary_YYYYMMDD_HHMMSS.xlsx
    """
    if not os.path.isfile(path):
        return {"success": False, "error": "Source file not found"}
    p = Path(path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"{p.stem}_summary_{ts}{p.suffix}"
    output_path = p.parent / output_name
    try:
        import shutil
        shutil.copy2(path, str(output_path))
        return {"success": True, "output_path": str(output_path), "output_name": output_name}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── Analysis Operations ──────────────────────────────────────────────

def create_summary_sheet(path: str, source_sheet: str, group_by_column: str, value_column: str, group_by_col_type: str = "text") -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if source_sheet not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{source_sheet}' not found"}
        ws = wb[source_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return {"success": False, "error": "No data in source sheet"}
        headers = list(rows[0]) if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        grp_idx = _find_column_index(headers, group_by_column)
        val_idx = _find_column_index(headers, value_column)
        if grp_idx is None or val_idx is None:
            wb.close()
            return {"success": False, "error": f"Columns '{group_by_column}' or '{value_column}' not found"}

        summary_data: dict[str, list[float]] = {}
        for row in data_rows:
            if grp_idx >= len(row) or val_idx >= len(row):
                continue
            key = str(row[grp_idx]) if row[grp_idx] is not None else "Unknown"
            try:
                val = float(row[val_idx]) if row[val_idx] is not None else 0.0
            except (ValueError, TypeError):
                val = 0.0
            if key not in summary_data:
                summary_data[key] = []
            summary_data[key].append(val)

        summary_name = f"{source_sheet[:20]}_Summary"
        if summary_name in wb.sheetnames:
            del wb[summary_name]
        ws_summary = wb.create_sheet(title=summary_name)
        ws_summary.append([group_by_column, "Count", f"Total {value_column}", f"Average {value_column}"])
        for cell in ws_summary[1]:
            cell.font = _style_header_font()
            cell.fill = _style_header_fill()
        for group, values in sorted(summary_data.items()):
            ws_summary.append([group, len(values), sum(values), sum(values) / len(values)])
        wb.save(path)
        wb.close()
        val_col_letter = get_column_letter(val_idx + 1) if val_idx is not None else "C"
        return {
            "success": True,
            "summary_sheet": summary_name,
            "group_count": len(summary_data),
            "total_rows": sum(len(v) for v in summary_data.values()),
            "value_column_letter": val_col_letter,
            **backup,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _find_column_index(headers: list, name: str) -> int | None:
    name_lower = name.strip().lower()
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == name_lower:
            return i
    return None


def create_pivot_table(path: str, source_sheet: str, group_by_column: str, value_column: str, value_agg: str = "sum") -> dict:
    return create_summary_sheet(path, source_sheet, group_by_column, value_column)


def compare_workbooks(path_a: str, path_b: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    info_a = open_workbook(path_a)
    info_b = open_workbook(path_b)
    if not info_a.get("success"):
        return {"success": False, "error": f"Cannot read {path_a}: {info_a.get('error')}"}
    if not info_b.get("success"):
        return {"success": False, "error": f"Cannot read {path_b}: {info_b.get('error')}"}

    common_sheets = set(info_a.get("sheet_names", [])) & set(info_b.get("sheet_names", []))
    if not common_sheets:
        return {"success": False, "error": "No common sheet names found"}

    differences = []
    for sheet in common_sheets:
        data_a = read_sheet(path_a, sheet)
        data_b = read_sheet(path_b, sheet)
        rows_a = data_a.get("rows", [])
        rows_b = data_b.get("rows", [])
        max_rows = max(len(rows_a), len(rows_b))
        row_diffs = 0
        for i in range(max_rows):
            if i >= len(rows_a) or i >= len(rows_b):
                row_diffs += 1
            elif rows_a[i] != rows_b[i]:
                row_diffs += 1
        differences.append({
            "sheet": sheet,
            "rows_a": len(rows_a),
            "rows_b": len(rows_b),
            "differing_rows": row_diffs,
            "identical": row_diffs == 0 and len(rows_a) == len(rows_b),
        })

    return {
        "success": True,
        "workbook_a": os.path.abspath(path_a),
        "workbook_b": os.path.abspath(path_b),
        "common_sheets": list(common_sheets),
        "sheet_differences": differences,
        "total_differences": sum(d["differing_rows"] for d in differences),
    }


def clean_csv(path: str, output_path: str | None = None) -> dict:
    if not os.path.isfile(path):
        return {"success": False, "error": "File not found"}
    try:
        if HAS_PANDAS:
            df = pd.read_csv(path)
            df = df.dropna(how="all")
            df = df.dropna(axis=1, how="all")
            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
            dest = output_path or str(Path(path).with_suffix(".xlsx"))
            with pd.ExcelWriter(dest, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Cleaned Data")
            return {"success": True, "output_path": os.path.abspath(dest), "rows_removed": 0, "columns_removed": 0, "original_rows": len(df)}
        else:
            import csv
            dest = output_path or str(Path(path).with_suffix(".xlsx"))
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cleaned Data"
            for row in rows:
                ws.append(row)
            wb.save(dest)
            wb.close()
            return {"success": True, "output_path": os.path.abspath(dest), "rows_removed": 0, "original_rows": len(rows)}
    except Exception as e:
        return {"success": False, "error": str(e)}


def split_by_category(path: str, source_sheet: str, category_column: str) -> dict:
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl not installed"}
    if not _is_path_allowed(path):
        return {"success": False, "error": "Path not in allowed workspace"}
    backup = create_backup_and_snapshot(path)
    try:
        wb = openpyxl.load_workbook(path)
        if source_sheet not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{source_sheet}' not found"}
        ws = wb[source_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return {"success": False, "error": "No data"}
        headers = list(rows[0])
        cat_idx = _find_column_index(headers, category_column)
        if cat_idx is None:
            wb.close()
            return {"success": False, "error": f"Column '{category_column}' not found"}

        categories: dict[str, list[list]] = {}
        for row in rows[1:]:
            cat = str(row[cat_idx]) if cat_idx < len(row) and row[cat_idx] is not None else "Uncategorized"
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(list(row))

        created_sheets = []
        for cat, cat_rows in sorted(categories.items()):
            safe_name = cat[:31]
            if safe_name in wb.sheetnames:
                del wb[safe_name]
            new_ws = wb.create_sheet(title=safe_name)
            new_ws.append(headers)
            for r in cat_rows:
                new_ws.append(r)
            created_sheets.append(safe_name)

        wb.save(path)
        wb.close()
        return {"success": True, "categories_found": len(created_sheets), "sheets_created": created_sheets, **backup}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── Google Sheets Placeholder ────────────────────────────────────────

def _google_sheets_unavailable():
    return {"success": False, "configured": False, "error": "Google Sheets integration not configured. Set up Google OAuth credentials to enable."}


def google_sheets_create_sheet(title: str) -> dict:
    return _google_sheets_unavailable()


def google_sheets_read_sheet(spreadsheet_id: str, range_name: str) -> dict:
    return _google_sheets_unavailable()


def google_sheets_write_range(spreadsheet_id: str, range_name: str, values: list[list[Any]]) -> dict:
    return _google_sheets_unavailable()


def google_sheets_apply_formula(spreadsheet_id: str, range_name: str, formula: str) -> dict:
    return _google_sheets_unavailable()


def google_sheets_export_xlsx(spreadsheet_id: str, output_path: str) -> dict:
    return _google_sheets_unavailable()
