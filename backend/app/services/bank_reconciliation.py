from __future__ import annotations

import csv
import io
import json
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("officepilot.bank_reconciliation")

MOCK_MODE = os.environ.get("BANK_FEED_MODE", "mock").lower() == "mock"


@dataclass
class BankTransaction:
    date: str
    description: str
    amount: float
    txn_type: str


@dataclass
class ReconciliationRecord:
    bank_date: str
    bank_description: str
    bank_amount: float
    bank_type: str
    matched_vendor: str | None
    invoice_amount: float | None
    invoice_id: str | None
    confidence: float
    status: str


class BankFeedAdapter:
    def parse_feed(self, file_path: str) -> list[BankTransaction]:
        if MOCK_MODE:
            return self._parse_mock(file_path)
        return self._parse_mock(file_path)

    def parse_feed_text(self, content: str, filename: str = "feed.csv") -> list[BankTransaction]:
        if filename.endswith(".csv"):
            return self._parse_csv_text(content)
        try:
            data = json.loads(content)
            return self._parse_json_data(data)
        except json.JSONDecodeError:
            return self._parse_csv_text(content)

    def _parse_mock(self, file_path: str) -> list[BankTransaction]:
        path = Path(file_path)
        if path.suffix.lower() == ".csv":
            return self._parse_csv_path(path)
        try:
            with open(file_path, encoding="utf-8") as f:
                data = json.load(f)
            return self._parse_json_data(data)
        except (json.JSONDecodeError, FileNotFoundError):
            return []

    def _parse_csv_path(self, path: Path) -> list[BankTransaction]:
        try:
            text = path.read_text(encoding="utf-8")
        except Exception:
            return []
        return self._parse_csv_text(text)

    def _parse_csv_text(self, text: str) -> list[BankTransaction]:
        transactions = []
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            row = {k.strip().lower(): v.strip() for k, v in row.items()}
            date = row.get("date", "")
            description = row.get("description", row.get("desc", ""))
            amount = self._parse_amount(row.get("amount", "0"))
            txn_type = row.get("type", row.get("txn_type", "debit")).lower()
            if date and description:
                transactions.append(BankTransaction(date, description, amount, txn_type))
        return transactions

    def _parse_json_data(self, data: list | dict) -> list[BankTransaction]:
        transactions = []
        items = data if isinstance(data, list) else data.get("transactions", [])
        for item in items:
            date = item.get("date", "")
            description = item.get("description", item.get("desc", ""))
            amount = self._parse_amount(item.get("amount", "0"))
            txn_type = item.get("type", item.get("txn_type", "debit")).lower()
            if date and description:
                transactions.append(BankTransaction(date, description, amount, txn_type))
        return transactions

    @staticmethod
    def _parse_amount(val: str | float) -> float:
        if isinstance(val, (int, float)):
            return float(val)
        cleaned = val.replace("$", "").replace(",", "").replace(" ", "")
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


class ReconciliationEngine:
    def __init__(self, semantic_memory: Any = None):
        self._semantic_memory = semantic_memory

    def _get_semantic_memory(self):
        if self._semantic_memory is not None:
            return self._semantic_memory
        from .semantic_memory import get_semantic_memory
        return get_semantic_memory()

    def reconcile(
        self,
        bank_transactions: list[BankTransaction],
        extracted_invoices: list[dict] | None = None,
        user_id: int | None = None,
    ) -> list[ReconciliationRecord]:
        records: list[ReconciliationRecord] = []
        sm = self._get_semantic_memory()

        for txn in bank_transactions:
            search_results = sm.semantic_search(
                query=txn.description,
                top_k=1,
                user_id=user_id,
            )

            if search_results:
                best = search_results[0]
                score = max(0.0, min(1.0, best.get("score", 0.0)))
                meta = best.get("metadata", {})
                matched_vendor = meta.get("vendor", meta.get("vendor_name", ""))
                invoice_amount = self._parse_float(meta.get("total_amount", meta.get("amount", 0)))
                invoice_id = best.get("id", "")

                if score >= 0.8:
                    status = "matched"
                elif score >= 0.5:
                    status = "fuzzy_match"
                else:
                    status = "unmatched"
            else:
                score = 0.0
                matched_vendor = ""
                invoice_amount = 0.0
                invoice_id = ""
                status = "unmatched"

            if extracted_invoices and status == "unmatched":
                best_exact = None
                best_confidence = 0.0
                for inv in extracted_invoices:
                    conf = self._exact_match_confidence(txn, inv)
                    if conf > best_confidence:
                        best_confidence = conf
                        best_exact = inv
                if best_exact and best_confidence > 0.0:
                    matched_vendor = best_exact.get("vendor", best_exact.get("vendor_name", matched_vendor))
                    invoice_amount = self._parse_float(best_exact.get("total_amount", best_exact.get("amount", 0)))
                    invoice_id = best_exact.get("id", best_exact.get("invoice_id", ""))
                    status = "fuzzy_match" if best_confidence >= 0.5 else "unmatched"
                    score = best_confidence

            records.append(ReconciliationRecord(
                bank_date=txn.date,
                bank_description=txn.description,
                bank_amount=txn.amount,
                bank_type=txn.txn_type,
                matched_vendor=matched_vendor or None,
                invoice_amount=invoice_amount or None,
                invoice_id=invoice_id or None,
                confidence=round(score, 4),
                status=status,
            ))

        return records

    @staticmethod
    def _exact_match_confidence(txn: BankTransaction, invoice: dict) -> float:
        description_words = set(txn.description.lower().split())
        vendor_word = str(invoice.get("vendor", invoice.get("vendor_name", ""))).lower()
        if not vendor_word:
            return 0.0
        vendor_parts = set(vendor_word.split())
        overlap = description_words & vendor_parts
        if not vendor_parts:
            return 0.0
        if abs(txn.amount) == abs(float(invoice.get("total_amount", invoice.get("amount", 0)))):
            return 1.0
        return len(overlap) / len(vendor_parts)

    @staticmethod
    def _parse_float(val: Any) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            return 0.0


def generate_reconciliation_excel(
    records: list[ReconciliationRecord],
    output_path: str,
) -> dict:
    from .excel_com_automation import ExcelComAdapter

    adapter = ExcelComAdapter(visible=False)
    with adapter as com:
        if not com.available:
            import openpyxl
            from openpyxl.styles import PatternFill, Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reconciliation"
            headers = ["Bank Date", "Bank Description", "Bank Amount", "Type", "Matched Vendor", "Invoice Amount", "Confidence", "Status"]
            ws.append(headers)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for rec in records:
                row = [rec.bank_date, rec.bank_description, rec.bank_amount, rec.bank_type, rec.matched_vendor or "", rec.invoice_amount or "", rec.confidence, rec.status]
                ws.append(row)
                row_num = ws.max_row
                cell = ws.cell(row=row_num, column=8)
                if rec.status == "matched":
                    cell.fill = green_fill
                    cell.font = Font(color="006100")
                elif rec.status == "fuzzy_match":
                    cell.fill = yellow_fill
                    cell.font = Font(color="9C6500")
                else:
                    cell.fill = red_fill
                    cell.font = Font(color="9C0006")
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 16
            ws.column_dimensions["G"].width = 12
            ws.column_dimensions["H"].width = 14
            wb.save(output_path)
            return {"status": "ok", "message": f"Reconciliation saved to {output_path}", "records": len(records)}

        try:
            workbook = com._app.books.add()
            sheet = workbook.sheets[0]
            sheet.name = "Reconciliation"
            headers = ["Bank Date", "Bank Description", "Bank Amount", "Type", "Matched Vendor", "Invoice Amount", "Confidence", "Status"]
            sheet.range("A1").value = [headers]
            for i, rec in enumerate(records, start=2):
                sheet.range(f"A{i}").value = [rec.bank_date, rec.bank_description, rec.bank_amount, rec.bank_type, rec.matched_vendor or "", rec.invoice_amount or "", rec.confidence, rec.status]

            header_range = sheet.range("A1:H1")
            header_range.api.Font.Bold = True
            header_range.api.Interior.Color = 0x4472C4
            header_range.api.Font.Color = 0xFFFFFF

            data_range = sheet.range(f"A2:H{len(records) + 1}")
            status_range = sheet.range(f"H2:H{len(records) + 1}")

            status_range.api.FormatConditions.Delete()
            fc_green = status_range.api.FormatConditions.Add(Type=2, Formula1='=H2="matched"')
            fc_green.Interior.Color = 0xC6EFCE
            fc_green.Font.Color = 0x006100
            fc_yellow = status_range.api.FormatConditions.Add(Type=2, Formula1='=H2="fuzzy_match"')
            fc_yellow.Interior.Color = 0xFFEB9C
            fc_yellow.Font.Color = 0x9C6500
            fc_red = status_range.api.FormatConditions.Add(Type=2, Formula1='=H2="unmatched"')
            fc_red.Interior.Color = 0xFFC7CE
            fc_red.Font.Color = 0x9C0006

            data_range.columns[0].column_width = 14
            data_range.columns[1].column_width = 40
            data_range.columns[2].column_width = 14
            data_range.columns[3].column_width = 10
            data_range.columns[4].column_width = 20
            data_range.columns[5].column_width = 16
            data_range.columns[6].column_width = 12
            data_range.columns[7].column_width = 14

            workbook.save(output_path)
            workbook.close()
            return {"status": "ok", "message": f"Reconciliation report saved to {output_path}", "records": len(records), "com": True}
        except Exception as e:
            logger.exception("COM Excel reconciliation failed")
            return {"status": "error", "message": f"COM error: {e}", "records": 0}
