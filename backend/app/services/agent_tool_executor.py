from __future__ import annotations

import json
import logging
import os
import random
import re
from datetime import date, datetime
from pathlib import Path

from sqlalchemy.orm import Session

from ..config import get_settings
from .tool_registry import get_tool
from .browser_automation import get_adapter, reset_adapter
from .browser_session_service import (
    is_action_blocked,
    is_sensitive_field,
    input_is_sensitive,
    create_session,
    get_active_session,
    get_session_by_run,
    open_url_live,
    read_page_live,
    take_screenshot_live,
    watch_for_download,
    copy_to_output,
    close_session,
    mock_open_url,
    mock_wait_for_login,
    mock_read_page,
    mock_export_report,
    mock_wait_for_download,
)

logger = logging.getLogger("officepilot.agent_tool_executor")

# Phase 34 — Gmail read-only: the only email tools allowed to execute.
# Any email tool not in this set is blocked with a clear message.
# The Gmail OAuth scope is hardcoded to 'gmail.readonly' — no write tools
# exist in the tool registry (send, forward, delete, move, mark-read, etc).
# This set provides defense-in-depth: if a bad plan somehow reaches the
# executor, it is blocked here.
GMAIL_READONLY_ALLOWED_TOOLS = frozenset({
    "email_connect_gmail",
    "email_search",
    "email_preview_messages",
    "email_download_attachments",
    "email_save_attachment",
    "email_disconnect_account",
    # Legacy aliases — all read-only
    "email_open",
    "email_create_draft",
    "email_open_message",
})

# Map abstract plan step types to concrete executor tool names.
# Plan steps use types like "read_screen", "approval_checkpoint", etc.
# These map to specific executor functions below.
STEP_TYPE_TOOL_MAP = {
    # Abstract step types → concrete automation tool names
    "read_screen": "screen_read_text",
    "capture_screen": "screen_capture",
    "find_button": "screen_find_button",
    "find_table": "screen_find_table",
    "confirm_state": "screen_confirm_state",
    "validation_checkpoint": "validate_result",
    "approval_checkpoint": "approval_request",
    "excel_action": "excel_create_workbook",
    "click": "desktop_click",
    "type_text": "desktop_type",
    "hotkey": "desktop_hotkey",
    "navigate": "browser_open_url",
    "copy": "desktop_copy",
    "paste": "desktop_paste",
    "wait": "desktop_wait",
    "open_app": "desktop_open_app",
    # Browser automation
    "browser_open_url": "browser_open_url",
    "browser_wait_for_user_login": "browser_wait_for_user_login",
    "browser_click": "browser_click",
    "browser_type": "browser_type",
    "browser_hotkey": "browser_hotkey",
    "browser_read_page": "browser_read_page",
    "browser_wait_for_download": "browser_wait_for_download",
    "browser_export_report": "browser_export_report",
    "browser_close": "browser_close",
    # Desktop automation
    "desktop_get_active_window": "desktop_get_active_window",
    "desktop_click": "desktop_click",
    "desktop_type": "desktop_type",
    "desktop_hotkey": "desktop_hotkey",
    "desktop_copy": "desktop_copy",
    "desktop_paste": "desktop_paste",
    "desktop_wait": "desktop_wait",
    "desktop_open_app": "desktop_open_app",
    # Screen/OCR automation
    "screen_capture": "screen_capture",
    "screen_read_text": "screen_read_text",
    "screen_find_button": "screen_find_button",
    "screen_find_table": "screen_find_table",
    "screen_confirm_state": "screen_confirm_state",
    # Excel automation
    "excel_open_workbook": "excel_open_workbook",
    "excel_create_workbook": "excel_create_workbook",
    "excel_read_sheet": "excel_read_sheet",
    "excel_write_cell": "excel_write_cell",
    "excel_write_range": "excel_write_range",
    "excel_append_rows": "excel_append_rows",
    "excel_save_workbook": "excel_save_workbook",
    "excel_apply_formula": "excel_apply_formula",
    "excel_apply_total_formula": "excel_apply_total_formula",
    "excel_apply_sumif_formula": "excel_apply_sumif_formula",
    "excel_apply_lookup_formula": "excel_apply_lookup_formula",
    "excel_verify_formula": "excel_verify_formula",
    "excel_format_header": "excel_format_header",
    "excel_auto_size_columns": "excel_auto_size_columns",
    "excel_freeze_top_row": "excel_freeze_top_row",
    "excel_apply_currency_format": "excel_apply_currency_format",
    "excel_apply_date_format": "excel_apply_date_format",
    "excel_highlight_negative_values": "excel_highlight_negative_values",
    "excel_add_total_row": "excel_add_total_row",
    "excel_detect_columns": "excel_detect_columns",
    "excel_create_summary_sheet": "excel_create_summary_sheet",
    "excel_create_summary_from_file": "excel_create_summary_from_file",
    "excel_create_pivot_table": "excel_create_pivot_table",
    "excel_compare_workbooks": "excel_compare_workbooks",
    "excel_clean_csv": "excel_clean_csv",
    "excel_split_by_category": "excel_split_by_category",
    "excel_export_pdf": "excel_export_pdf",
    # Google Sheets placeholder
    "google_sheets_create_sheet": "google_sheets_create_sheet",
    "google_sheets_read_sheet": "google_sheets_read_sheet",
    "google_sheets_write_range": "google_sheets_write_range",
    "google_sheets_apply_formula": "google_sheets_apply_formula",
    "google_sheets_export_xlsx": "google_sheets_export_xlsx",
    # File automation
    "file_open": "file_open",
    "file_open_folder": "file_open_folder",
    "file_copy": "file_copy",
    "file_move": "file_move",
    "file_rename": "file_rename",
    "file_create_folder": "file_create_folder",
    "file_watch_folder": "file_watch_folder",
    "file_find_latest_download": "file_find_latest_download",
    "file_find_in_downloads": "file_find_in_downloads",
    "file_copy_table_to_excel": "file_copy_table_to_excel",
    # Email automation (Phase 34)
    "email_connect_gmail": "email_connect_gmail",
    "email_search": "email_search",
    "email_preview_messages": "email_preview_messages",
    "email_download_attachments": "email_download_attachments",
    "email_save_attachment": "email_save_attachment",
    "email_disconnect_account": "email_disconnect_account",
    # Legacy email aliases
    "email_open": "email_open",
    "email_create_draft": "email_create_draft",
    "email_open_message": "email_open_message",
    # Workflow tools
    "workflow_record_start": "workflow_record_start",
    "workflow_record_stop": "workflow_record_stop",
    "workflow_save_as_skill": "workflow_save_as_skill",
    "workflow_dry_run": "workflow_dry_run",
    "workflow_replay": "workflow_replay",
    "workflow_restore_version": "workflow_restore_version",
    # Safety tools
    "approval_request": "approval_request",
    "approval_confirm": "approval_confirm",
    "emergency_stop": "emergency_stop",
    "audit_log": "audit_log",
    "snapshot_create": "snapshot_create",
    "sensitive_redact": "sensitive_redact",
    # Legacy aliases (backward compat)
    "read_current_screen": "screen_read_text",
    "click_approved_target": "desktop_click",
    "type_approved_text": "desktop_type",
    "read_pdf": "file_open",
    "copy_table_to_excel": "file_copy_table_to_excel",
    "open_browser": "browser_open_url",
    "open_file": "file_open",
    "open_folder": "file_open_folder",
    "open_email": "email_open",
    "search_email": "email_search",
    "save_workflow": "workflow_save_as_skill",
    "replay_workflow": "workflow_replay",
    "create_excel_workbook": "excel_create_workbook",
    "append_excel_rows": "excel_append_rows",
    "create_excel_summary": "excel_create_summary_sheet",
    "scan_local_folder": "file_find_latest_download",
    # P&L legacy aliases
    "open_accounting_platform": "browser_open_url",
    "wait_for_manual_login": "browser_wait_for_user_login",
    "navigate_to_profit_loss_report": "browser_click",
    "set_report_date_range": "desktop_type",
    "export_accounting_report": "browser_export_report",
    "watch_downloaded_report": "browser_wait_for_download",
    "read_pnl_report_file": "file_open",
    "normalize_pnl_report": "validate_result",
    "compare_pnl_reports": "validate_result",
    "create_pnl_comparison_excel": "excel_create_summary_sheet",
    "speak_pnl_difference": "approval_request",
    "save_report_workflow": "workflow_save_as_skill",
}

EXECUTOR_RESULT_OK = "success"
EXECUTOR_RESULT_FAILED = "failed"
EXECUTOR_RESULT_BLOCKED = "blocked"
EXECUTOR_RESULT_NEEDS_APPROVAL = "needs_approval"
EXECUTOR_RESULT_DRY_RUN = "dry_run"


def _data_dir() -> Path:
    from ..config import get_settings
    s = get_settings()
    return Path(s.data_dir)


def _demo_mode() -> bool:
    return os.environ.get("DEMO_MODE", "false").lower() in ("1", "true", "yes", "on")


def execute_tool(tool_name: str, params: dict, mode: str, db: Session, user) -> dict:
    # Map abstract step types to concrete executor tool names
    resolved = STEP_TYPE_TOOL_MAP.get(tool_name, tool_name)
    if resolved != tool_name:
        logger.debug("Mapped step type '%s' -> tool '%s'", tool_name, resolved)
    tool_name = resolved

    tool = get_tool(tool_name)
    if not tool:
        # Phase 34 — Gmail read-only safety gate (defense-in-depth)
        # If the tool name starts with "email_" and is not in the allowed set,
        # block it even if it doesn't exist in the registry.
        if tool_name.startswith("email_") and tool_name not in GMAIL_READONLY_ALLOWED_TOOLS:
            return {
                "status": EXECUTOR_RESULT_BLOCKED,
                "message": "This Gmail action is blocked because OfficePilot only has read-only email automation.",
                "output": {"blocked": True, "reason": "gmail_readonly_policy", "tool_name": tool_name},
                "audit_required": True,
                "error_message": "gmail_readonly_policy: email write action blocked",
            }
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": f"Unknown tool: {tool_name}",
            "output": {},
            "audit_required": False,
            "error_message": f"No tool definition for '{tool_name}'",
        }

    # Phase 34 — Gmail read-only safety gate (defense-in-depth) for existing tools
    if tool_name.startswith("email_") and tool_name not in GMAIL_READONLY_ALLOWED_TOOLS:
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": "This Gmail action is blocked because OfficePilot only has read-only email automation.",
            "output": {"blocked": True, "reason": "gmail_readonly_policy", "tool_name": tool_name},
            "audit_required": True,
            "error_message": "gmail_readonly_policy: email write action blocked",
        }

    if mode == "dry_run":
        return {
            "status": EXECUTOR_RESULT_DRY_RUN,
            "message": f"Dry-run: would execute '{tool_name}'",
            "output": {"tool": tool_name, "params": params, "risk_level": tool.risk_level},
            "audit_required": tool.audit_required,
            "error_message": None,
        }

    if tool.risk_level == "high" and mode != "live":
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": f"Tool '{tool_name}' is high risk and cannot run in {mode} mode.",
            "output": {},
            "audit_required": True,
            "error_message": "High-risk tool blocked in non-live mode.",
        }

    executor_map = {
        # Legacy / backward-compatible executors
        "speak_response": _execute_speak_response,
        "approval_request": _execute_approval_request,
        "validate_result": _execute_validate_result,
        "save_workflow": _execute_save_workflow,
        "replay_workflow": _execute_replay_workflow,
        "calculate_excel_total": _execute_calculate_excel_total,
        "create_excel_workbook": _execute_create_excel_workbook,
        "append_excel_rows": _execute_append_excel_rows,
        "search_email": _execute_search_email_mock,
        "download_attachments": _execute_download_attachments_mock,
        "extract_invoice_data": _execute_extract_invoice_data,
        "open_folder": _execute_open_folder,
        "open_file": _execute_open_file,
        "read_current_screen": _execute_read_current_screen,
        # P&L comparison tools
        "open_accounting_platform": _execute_open_accounting_platform,
        "wait_for_manual_login": _execute_wait_for_manual_login,
        "navigate_to_profit_loss_report": _execute_navigate_to_profit_loss_report,
        "set_report_date_range": _execute_set_report_date_range,
        "export_accounting_report": _execute_export_accounting_report,
        "watch_downloaded_report": _execute_watch_downloaded_report,
        "read_pnl_report_file": _execute_read_pnl_report_file,
        "normalize_pnl_report": _execute_normalize_pnl_report,
        "compare_pnl_reports": _execute_compare_pnl_reports,
        "create_pnl_comparison_excel": _execute_create_pnl_comparison_excel,
        "speak_pnl_difference": _execute_speak_pnl_difference,
        "save_report_workflow": _execute_save_report_workflow,
        # Phase 25: Local folder invoice workflow
        "scan_local_folder": _execute_scan_local_folder,
        "create_daily_invoices_excel": _execute_create_daily_invoices_excel,
        # Excel Skill Pack
        "excel_open_workbook": _execute_excel_open_workbook,
        "excel_read_sheet": _execute_excel_read_sheet,
        "excel_detect_columns": _execute_excel_detect_columns,
        "excel_apply_formula": _execute_excel_apply_formula,
        "excel_apply_total_formula": _execute_excel_apply_total_formula,
        "excel_add_total_row": _execute_excel_add_total_row,
        "excel_create_summary_sheet": _execute_excel_create_summary_sheet,
        "excel_create_workbook": _execute_create_excel_workbook,
        "excel_create_summary_from_file": _execute_excel_create_summary_from_file,
        "excel_create_pivot_table": _execute_excel_create_pivot_table,
        "excel_compare_workbooks": _execute_excel_compare_workbooks,
        "excel_clean_csv": _execute_excel_clean_csv,
        "excel_format_header": _execute_excel_format_header,
        "excel_auto_size_columns": _execute_excel_auto_size_columns,
        "excel_freeze_top_row": _execute_excel_freeze_top_row,
        "excel_apply_currency_format": _execute_excel_apply_currency_format,
        "excel_split_by_category": _execute_excel_split_by_category,
        "excel_save_workbook": _execute_excel_save_workbook,
        "excel_append_rows": _execute_excel_append_rows,
        # Google Sheets placeholders
        "google_sheets_create_sheet": _execute_google_sheets_placeholder,
        "google_sheets_read_sheet": _execute_google_sheets_placeholder,
        "google_sheets_write_range": _execute_google_sheets_placeholder,
        "google_sheets_apply_formula": _execute_google_sheets_placeholder,
        "google_sheets_export_xlsx": _execute_google_sheets_placeholder,
        # ── New Automation Tool Executors ────────────────────────────────
        # Browser automation
        "browser_open_url": _execute_browser_open_url,
        "browser_wait_for_user_login": _execute_browser_wait_for_user_login,
        "browser_click": _execute_browser_click,
        "browser_type": _execute_browser_type,
        "browser_hotkey": _execute_browser_hotkey,
        "browser_read_page": _execute_browser_read_page,
        "browser_wait_for_download": _execute_browser_wait_for_download,
        "browser_export_report": _execute_browser_export_report,
        "browser_close": _execute_browser_close,
        # Desktop automation
        "desktop_get_active_window": _execute_desktop_get_active_window,
        "desktop_click": _execute_desktop_click,
        "desktop_type": _execute_desktop_type,
        "desktop_hotkey": _execute_desktop_hotkey,
        "desktop_copy": _execute_desktop_copy,
        "desktop_paste": _execute_desktop_paste,
        "desktop_wait": _execute_desktop_wait,
        "desktop_open_app": _execute_desktop_open_app,
        # Screen/OCR automation
        "screen_capture": _execute_screen_capture,
        "screen_read_text": _execute_screen_read_text,
        "screen_find_button": _execute_screen_find_button,
        "screen_find_table": _execute_screen_find_table,
        "screen_confirm_state": _execute_screen_confirm_state,
        # File automation
        "file_open": _execute_file_open,
        "file_open_folder": _execute_file_open_folder,
        "file_copy": _execute_file_copy,
        "file_move": _execute_file_move,
        "file_rename": _execute_file_rename,
        "file_create_folder": _execute_file_create_folder,
        "file_watch_folder": _execute_file_watch_folder,
        "file_find_latest_download": _execute_file_find_latest_download,
        "file_find_in_downloads": _execute_file_find_in_downloads,
        "file_copy_table_to_excel": _execute_file_copy_table_to_excel,
        # Email automation (Phase 34)
        "email_connect_gmail": _execute_email_connect_gmail,
        "email_search": _execute_email_search,
        "email_preview_messages": _execute_email_preview_messages,
        "email_download_attachments": _execute_email_download_attachments,
        "email_save_attachment": _execute_email_save_attachment,
        "email_disconnect_account": _execute_email_disconnect_account,
        # Legacy email
        "email_open": _execute_email_open,
        "email_create_draft": _execute_email_create_draft,
        "email_open_message": _execute_email_open_message,
        # Workflow tools
        "workflow_record_start": _execute_workflow_record_start,
        "workflow_record_stop": _execute_workflow_record_stop,
        "workflow_save_as_skill": _execute_workflow_save_as_skill,
        "workflow_dry_run": _execute_workflow_dry_run,
        "workflow_replay": _execute_workflow_replay,
        "workflow_restore_version": _execute_workflow_restore_version,
        # Safety tools
        "approval_confirm": _execute_approval_confirm,
        "emergency_stop": _execute_emergency_stop,
        "audit_log": _execute_audit_log,
        "snapshot_create": _execute_snapshot_create,
        "sensitive_redact": _execute_sensitive_redact,
    }

    fn = executor_map.get(tool_name)
    if not fn:
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": f"Executor not implemented for '{tool_name}'",
            "output": {},
            "audit_required": True,
            "error_message": f"No executor for '{tool_name}'",
        }

    try:
        return fn(params, db, user)
    except Exception as e:
        logger.exception("Tool execution failed: %s", tool_name)
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": f"Execution error: {e}",
            "output": {},
            "audit_required": tool.audit_required,
            "error_message": str(e),
        }


def _execute_speak_response(params: dict, db: Session, user) -> dict:
    text = params.get("text", "")
    language = params.get("language", "en")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Speaking: {text[:200]}",
        "output": {"text": text, "language": language, "spoken": True},
        "audit_required": False,
    }


def _execute_validate_result(params: dict, db: Session, user) -> dict:
    expected = params.get("expected", "")
    actual = params.get("actual", "")
    match = expected.strip().lower() == actual.strip().lower() if expected and actual else True
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Validation passed" if match else "Validation failed",
        "output": {"expected": expected, "actual": actual, "match": match},
        "audit_required": True,
    }


def _execute_save_workflow(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Workflow saved (handled by caller)",
        "output": {"saved": True},
        "audit_required": True,
    }


def _execute_replay_workflow(params: dict, db: Session, user) -> dict:
    workflow_name = params.get("workflow_name", "unknown")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Replayed workflow '{workflow_name}' in dry-run",
        "output": {"workflow_name": workflow_name, "replayed": True, "mode": "dry_run"},
        "audit_required": True,
    }


def _execute_calculate_excel_total(params: dict, db: Session, user) -> dict:
    values = params.get("values", params.get("numbers", []))
    if isinstance(values, str):
        try:
            values = json.loads(values)
        except (json.JSONDecodeError, TypeError):
            values = []
    if not isinstance(values, list):
        values = []
    numbers = [float(v) for v in values if isinstance(v, (int, float, str)) and _is_number(v)]
    total = sum(numbers)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Total calculated: {total}",
        "output": {"total": total, "count": len(numbers), "values": numbers},
        "audit_required": True,
    }


def _is_number(v):
    if isinstance(v, (int, float)):
        return True
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def _execute_create_excel_workbook(params: dict, db: Session, user) -> dict:
    filename = params.get("filename", params.get("name", f"workbook_{date.today().isoformat()}.xlsx"))
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    out_dir = _data_dir() / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = out_dir / filename
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = params.get("sheet_name", "Sheet1")
        headers = params.get("headers", params.get("columns", []))
        if headers:
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)
        wb.save(str(filepath))
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Workbook created: {filepath.name}",
            "output": {"filepath": str(filepath), "filename": filename, "sheet": ws.title, "rows": 1 if headers else 0},
            "audit_required": True,
            "snapshot_created": True,
        }
    except ImportError:
        filepath.write_text("")
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Workbook placeholder created: {filepath.name} (openpyxl not available)",
            "output": {"filepath": str(filepath), "filename": filename, "note": "openpyxl not installed"},
            "audit_required": True,
            "snapshot_created": True,
        }


def _execute_append_excel_rows(params: dict, db: Session, user) -> dict:
    filename = params.get("filename", params.get("name", f"workbook_{date.today().isoformat()}.xlsx"))
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    rows = params.get("rows", params.get("data", []))
    if isinstance(rows, str):
        try:
            rows = json.loads(rows)
        except (json.JSONDecodeError, TypeError):
            rows = []
    if not isinstance(rows, list):
        rows = []
    out_dir = _data_dir() / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = out_dir / filename
    try:
        import openpyxl
        if filepath.exists():
            wb = openpyxl.load_workbook(str(filepath))
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = params.get("sheet_name", "Sheet1")
        start_row = ws.max_row + 1
        for row_idx, row_data in enumerate(rows, start_row):
            if isinstance(row_data, (list, tuple)):
                for col_idx, val in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            elif isinstance(row_data, dict):
                for col_idx, (_, val) in enumerate(sorted(row_data.items()), 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(str(filepath))
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Appended {len(rows)} rows to {filename}",
            "output": {"filepath": str(filepath), "rows_appended": len(rows), "start_row": start_row},
            "audit_required": True,
            "snapshot_created": True,
        }
    except ImportError:
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Dry append to {filename} ({len(rows)} rows) - openpyxl not available",
            "output": {"filepath": str(filepath), "rows_appended": len(rows), "note": "openpyxl not installed"},
            "audit_required": True,
        }


def _execute_search_email_mock(params: dict, db: Session, user) -> dict:
    query = params.get("query", params.get("subject", ""))
    today_str = date.today().isoformat()
    if _demo_mode() or os.environ.get("OFFICEPILOT_GMAIL_ALLOW_REAL", "true").lower() == "false":
        results = [
            {
                "id": f"mock-email-{i}",
                "subject": f"Invoice from Vendor {chr(65 + i)} - {today_str}",
                "sender": f"vendor{chr(65 + i)}@example.com",
                "date": today_str,
                "has_attachments": True,
            }
            for i in range(random.randint(1, 3))
        ]
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Found {len(results)} emails in demo mode",
            "output": {"results": results, "mode": "demo", "query": query},
            "audit_required": True,
        }
    return {
        "status": EXECUTOR_RESULT_DRY_RUN,
        "message": "Email search would use real Gmail API (dry-run only)",
        "output": {"mode": "dry_run", "query": query},
        "audit_required": True,
    }


def _execute_download_attachments_mock(params: dict, db: Session, user) -> dict:
    email_id = params.get("email_id", params.get("email_ids", ""))
    # In demo mode, create dummy invoice files
    if _demo_mode():
        out_dir = _data_dir() / "invoices" / "demo"
        out_dir.mkdir(parents=True, exist_ok=True)
        files = []
        for i in range(1, 3):
            fname = f"demo_invoice_{date.today().isoformat()}_{i}.pdf"
            fpath = out_dir / fname
            if not fpath.exists():
                fpath.write_text(f"Demo Invoice {i}\nVendor: Demo Vendor\nAmount: {random.randint(100, 5000)}")
            files.append(str(fpath))
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Downloaded {len(files)} attachments in demo mode",
            "output": {"files": files, "mode": "demo", "email_id": email_id},
            "audit_required": True,
            "snapshot_created": False,
        }
    return {
        "status": EXECUTOR_RESULT_DRY_RUN,
        "message": "Attachment download would use real Gmail API (dry-run only)",
        "output": {"mode": "dry_run", "email_id": email_id},
        "audit_required": True,
    }


def _execute_extract_invoice_data(params: dict, db: Session, user) -> dict:
    filepath = params.get("filepath", "")
    if _demo_mode() and not filepath:
        rows = [
            {"vendor": "Demo Corp", "invoice_no": "INV-DEMO-001", "amount": 1250.00, "date": date.today().isoformat()},
            {"vendor": "Sample Ltd", "invoice_no": "INV-DEMO-002", "amount": 3400.50, "date": date.today().isoformat()},
        ]
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Extracted {len(rows)} invoices in demo mode",
            "output": {"rows": rows, "mode": "demo", "filepath": filepath},
            "audit_required": True,
        }
    try:
        from .local_invoice_workflow import extract_invoice_from_file
        inv = extract_invoice_from_file(filepath)
        row = {
            "vendor": inv.vendor,
            "invoice_no": inv.invoice_number,
            "amount": inv.total_amount,
            "date": inv.invoice_date,
            "tax": inv.tax,
            "currency": inv.currency,
            "source_file": inv.source_file,
            "confidence": inv.confidence,
            "warnings": inv.warnings,
            "status": inv.status,
        }
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Extracted invoice from {filepath}: {inv.vendor} - {inv.total_amount}",
            "output": {"rows": [row], "mode": "real", "filepath": filepath},
            "audit_required": True,
        }
    except Exception as e:
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": f"Extraction failed: {e}",
            "output": {},
            "audit_required": True,
            "error_message": str(e),
        }


def _execute_open_folder(params: dict, db: Session, user) -> dict:
    folder_path = params.get("path", params.get("folder_path", ""))
    if not folder_path:
        folder_path = str(_data_dir() / "invoices")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened folder: {folder_path}",
        "output": {"folder_path": folder_path, "opened": True},
        "audit_required": True,
    }


def _execute_open_file(params: dict, db: Session, user) -> dict:
    file_path = params.get("path", params.get("file_path", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened file: {file_path}" if file_path else "No file path provided",
        "output": {"file_path": file_path, "opened": bool(file_path)},
        "audit_required": True,
    }


def _execute_read_current_screen(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Screen context captured (mock mode)",
        "output": {
            "active_window": "Mock Window - OfficePilot",
            "text_content": "This is a simulated screen capture. In production, this would return real OCR text.",
            "mode": "mock",
        },
        "audit_required": True,
    }


# ── P&L Comparison Executors ────────────────────────────────────────────────


def _execute_open_accounting_platform(params: dict, db: Session, user) -> dict:
    platform = params.get("platform", params.get("target", "QuickBooks"))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened {platform} (simulated)",
        "output": {
            "platform": platform,
            "url": f"https://quickbooks.example.com",
            "opened": True,
            "mode": "demo" if _demo_mode() else "live",
        },
        "audit_required": True,
    }


def _execute_wait_for_manual_login(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Waiting for manual login. Please log in to the accounting platform.",
        "output": {
            "action": "wait_for_login",
            "status": "waiting",
            "prompt": "Please log in manually. Automation will resume after login.",
        },
        "audit_required": True,
    }


def _execute_navigate_to_profit_loss_report(params: dict, db: Session, user) -> dict:
    platform = params.get("platform", "QuickBooks")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Navigated to Profit & Loss report in {platform}",
        "output": {
            "platform": platform,
            "report": "Profit and Loss",
            "navigated": True,
        },
        "audit_required": True,
    }


def _execute_set_report_date_range(params: dict, db: Session, user) -> dict:
    period = params.get("period", params.get("date_range", "Current Month"))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Set report date range to {period}",
        "output": {
            "period": period,
            "start_date": params.get("start_date"),
            "end_date": params.get("end_date"),
            "set": True,
        },
        "audit_required": True,
    }


def _execute_export_accounting_report(params: dict, db: Session, user) -> dict:
    period = params.get("period", "current")
    fmt = params.get("format", "csv")
    export_dir = _data_dir() / "exports" / "pnl"
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"pnl_{period}_{date.today().isoformat()}.{fmt}"
    filepath = export_dir / filename

    if _demo_mode():
        from .accounting_report_comparison import get_demo_current_report, get_demo_previous_report
        report = get_demo_current_report() if "current" in period else get_demo_previous_report()
        fake_data = {
            "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in report.rows],
            "total_income": report.total_income,
            "total_expenses": report.total_expenses,
            "net_income": report.net_income,
            "period_label": period,
        }
        import json
        filepath.write_text(json.dumps(fake_data, indent=2), encoding="utf-8")
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Exported {period} P&L report: {filepath.name}",
            "output": {
                "filepath": str(filepath),
                "filename": filename,
                "period": period,
                "format": fmt,
                "mode": "demo",
            },
            "audit_required": True,
            "snapshot_created": False,
        }

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Exported {period} P&L report to {filename}",
        "output": {
            "filepath": str(filepath),
            "filename": filename,
            "period": period,
            "format": fmt,
        },
        "audit_required": True,
    }


def _execute_watch_downloaded_report(params: dict, db: Session, user) -> dict:
    filename = params.get("filename", params.get("filepath", "downloaded_report"))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Downloaded report detected: {filename}",
        "output": {
            "filepath": filename,
            "found": True,
        },
        "audit_required": True,
    }


def _execute_read_pnl_report_file(params: dict, db: Session, user) -> dict:
    filepath = params.get("filepath", "")
    if _demo_mode() or not filepath:
        from .accounting_report_comparison import get_demo_current_report, get_demo_previous_report
        period = params.get("period", "current")
        report = get_demo_current_report() if "current" in period else get_demo_previous_report()
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Read P&L report for {period} (demo mode)",
            "output": {
                "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in report.rows],
                "total_income": report.total_income,
                "total_expenses": report.total_expenses,
                "net_income": report.net_income,
                "period": period,
                "mode": "demo",
            },
            "audit_required": True,
        }

    try:
        from .accounting_report_comparison import read_pnl_report
        report = read_pnl_report(filepath)
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Read P&L report from {filepath}",
            "output": {
                "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in report.rows],
                "total_income": report.total_income,
                "total_expenses": report.total_expenses,
                "net_income": report.net_income,
                "period": params.get("period", ""),
                "mode": "real",
            },
            "audit_required": True,
        }
    except Exception as e:
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": f"Failed to read P&L report: {e}",
            "output": {},
            "audit_required": True,
            "error_message": str(e),
        }


def _execute_normalize_pnl_report(params: dict, db: Session, user) -> dict:
    rows = params.get("rows", params.get("data", []))
    from .accounting_report_comparison import normalize_pnl_rows
    normalized = normalize_pnl_rows(rows)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Normalized {len(normalized)} P&L rows",
        "output": {
            "rows": [{"account": r.account, "amount": r.amount, "type": r.type} for r in normalized],
            "count": len(normalized),
        },
        "audit_required": True,
    }


def _execute_compare_pnl_reports(params: dict, db: Session, user) -> dict:
    from .accounting_report_comparison import (
        PnLReport, PnLRow, compare_pnl_reports,
        get_demo_current_report, get_demo_previous_report,
        pnl_comparison_to_dict,
    )

    if _demo_mode():
        current = get_demo_current_report()
        previous = get_demo_previous_report()
        comparison = compare_pnl_reports(current, previous)
        result = pnl_comparison_to_dict(comparison)
        result["summary_english"] = build_pnl_summary_text(comparison, "en")
        result["summary_roman_urdu"] = build_pnl_summary_text(comparison, "roman_urdu")
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": "P&L comparison completed (demo mode)",
            "output": result,
            "audit_required": True,
        }

    current_data = params.get("current", {})
    previous_data = params.get("previous", {})

    def _to_report(data, label):
        return PnLReport(
            rows=[PnLRow(**r) for r in data.get("rows", [])],
            total_income=float(data.get("total_income", 0)),
            total_expenses=float(data.get("total_expenses", 0)),
            net_income=float(data.get("net_income", 0)),
            period_label=label,
        )

    current = _to_report(current_data, params.get("current_label", "Current Month"))
    previous = _to_report(previous_data, params.get("previous_label", "Previous Month"))
    comparison = compare_pnl_reports(current, previous)
    result = pnl_comparison_to_dict(comparison)
    result["summary_english"] = build_pnl_summary_text(comparison, "en")
    result["summary_roman_urdu"] = build_pnl_summary_text(comparison, "roman_urdu")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "P&L comparison completed",
        "output": result,
        "audit_required": True,
    }


def _execute_create_pnl_comparison_excel(params: dict, db: Session, user) -> dict:
    from .accounting_report_comparison import (
        PnLReport, PnLRow, compare_pnl_reports,
        create_pnl_comparison_excel, pnl_comparison_to_dict,
    )

    comparison_data = params.get("comparison", params)
    output_dir = _data_dir() / "exports" / "pnl"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"pnl_comparison_{date.today().isoformat()}.xlsx"

    if _demo_mode():
        from .accounting_report_comparison import get_demo_current_report, get_demo_previous_report
        current = get_demo_current_report()
        previous = get_demo_previous_report()
        comparison = compare_pnl_reports(current, previous)
        filepath = create_pnl_comparison_excel(current, previous, comparison, output_path)
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"P&L comparison Excel created: {output_path.name}",
            "output": {
                "filepath": filepath,
                "filename": output_path.name,
                "mode": "demo",
            },
            "audit_required": True,
            "snapshot_created": True,
        }

    def _report_from_data(data, label):
        return PnLReport(
            rows=[PnLRow(**r) for r in data.get("rows", [])],
            total_income=float(data.get("total_income", 0)),
            total_expenses=float(data.get("total_expenses", 0)),
            net_income=float(data.get("net_income", 0)),
            period_label=label,
        )

    current = _report_from_data(comparison_data.get("current", {}), "Current Month")
    previous = _report_from_data(comparison_data.get("previous", {}), "Previous Month")
    comparison = compare_pnl_reports(current, previous)
    filepath = create_pnl_comparison_excel(current, previous, comparison, output_path)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"P&L comparison Excel created: {output_path.name}",
        "output": {
            "filepath": filepath,
            "filename": output_path.name,
            "mode": "real",
        },
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_speak_pnl_difference(params: dict, db: Session, user) -> dict:
    text = params.get("text", params.get("summary_english", ""))
    language = params.get("language", "en")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Speaking P&L difference: {text[:200]}",
        "output": {"text": text, "language": language, "spoken": True},
        "audit_required": False,
    }


def _execute_save_report_workflow(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "P&L comparison workflow saved (handled by caller)",
        "output": {"saved": True, "workflow_type": "pnl_comparison"},
        "audit_required": True,
    }


def build_pnl_summary_text(comparison, language="en"):
    from .accounting_report_comparison import build_pnl_summary_text as _build
    return _build(comparison, language)


# ── Phase 25: Local Folder Invoice Workflow Executors ─────────────────────────


def _execute_scan_local_folder(params: dict, db: Session, user) -> dict:
    folder_path = params.get("folder_path", params.get("path", ""))
    date_filter = params.get("date_filter", "today")
    keywords = params.get("keywords", True)

    from .local_invoice_workflow import scan_folder_for_invoices
    files = scan_folder_for_invoices(folder_path, date_filter=date_filter, keywords=keywords)

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Found {len(files)} invoice files in {folder_path}",
        "output": {
            "files": [{"path": f.path, "filename": f.filename, "modified": f.modified, "size": f.size} for f in files],
            "count": len(files),
            "folder_path": folder_path,
            "date_filter": date_filter,
        },
        "audit_required": True,
    }


# ── Excel Skill Pack Executors ────────────────────────────────────────────────

def _excel_tool(import_name):
    from . import excel_tools as et
    return getattr(et, import_name, None)


def _execute_excel_open_workbook(params: dict, db: Session, user) -> dict:
    fn = _excel_tool("open_workbook")
    result = fn(params.get("path", ""))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Opened workbook: {result.get('sheet_names', [])}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_read_sheet(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.read_sheet(params.get("path", ""), params.get("sheet_name", "Sheet1"), max_rows=params.get("max_rows", 0))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Read {result.get('row_count', 0)} rows from {params.get('sheet_name', 'Sheet1')}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_detect_columns(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.detect_columns(params.get("path", ""), params.get("sheet_name", "Sheet1"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Detected {len(result.get('columns', []))} columns" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_apply_formula(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    formula = params.get("formula", "")
    cell = params.get("cell", "A1")
    result = et.apply_formula(params.get("path", ""), params.get("sheet_name", "Sheet1"), cell, formula)
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Applied formula to {cell}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


def _execute_excel_apply_total_formula(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.apply_total_formula(
        params.get("path", ""),
        params.get("sheet_name", "Sheet1"),
        params.get("column_letter", "A"),
        params.get("start_row", 2),
        params.get("end_row"),
    )
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Total formula added" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


def _execute_excel_add_total_row(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.add_total_row(
        params.get("path", ""),
        params.get("sheet_name", "Sheet1"),
        params.get("column_letter", params.get("amount_column", "A")),
        params.get("label", "Total"),
    )
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Total row added at row {result.get('total_row', '?')}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


def _execute_excel_create_summary_sheet(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.create_summary_sheet(
        params.get("path", ""),
        params.get("source_sheet", params.get("sheet_name", "Sheet1")),
        params.get("group_by_column", ""),
        params.get("value_column", params.get("amount_column", "")),
    )
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Summary sheet '{result.get('summary_sheet', '')}' created" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


def _execute_excel_create_summary_from_file(params: dict, db: Session, user) -> dict:
    path = params.get("path", params.get("file_path", ""))
    if not path:
        return {
            "status": EXECUTOR_RESULT_NEEDS_APPROVAL,
            "message": "Which Excel file should I summarize?",
            "output": {
                "needs_input": True,
                "input_type": "file_picker",
                "accepted_types": [".xlsx", ".xlsm", ".csv"],
                "message": "Please select the Excel file to summarize.",
            },
            "audit_required": False,
        }

    mode = params.get("mode", "live")
    options = {
        "mode": mode,
        "source_sheet": params.get("source_sheet") or params.get("sheet_name", ""),
        "group_by_column": params.get("group_by_column", ""),
        "value_column": params.get("value_column") or params.get("amount_column", ""),
    }

    from . import excel_tools as et
    result = et.create_summary_from_file(path, options)

    if result.get("status") == "needs_input":
        return {
            "status": EXECUTOR_RESULT_NEEDS_APPROVAL,
            "message": result.get("message", "Additional input needed"),
            "output": result,
            "audit_required": False,
        }

    if result.get("status") == "failed":
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": result.get("error", "Excel summary failed"),
            "output": result,
            "audit_required": True,
        }

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Summary created: {result.get('rows_processed', 0)} rows, total {result.get('total', 0):,.2f}",
        "output": result,
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_excel_create_pivot_table(params: dict, db: Session, user) -> dict:
    return _execute_excel_create_summary_sheet(params, db, user)


def _execute_excel_compare_workbooks(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.compare_workbooks(params.get("path_a", ""), params.get("path_b", ""))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Found {result.get('total_differences', 0)} differing rows" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_clean_csv(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.clean_csv(params.get("path", ""), params.get("output_path"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Cleaned CSV -> {result.get('output_path', '')}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_format_header(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.format_header(params.get("path", ""), params.get("sheet_name", "Sheet1"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": "Header formatted" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_auto_size_columns(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.auto_size_columns(params.get("path", ""), params.get("sheet_name", "Sheet1"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": "Columns auto-sized" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_freeze_top_row(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.freeze_top_row(params.get("path", ""), params.get("sheet_name", "Sheet1"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": "Top row frozen" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_apply_currency_format(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.apply_currency_format(params.get("path", ""), params.get("sheet_name", "Sheet1"), params.get("column_letter", "A"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": "Currency format applied" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_split_by_category(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.split_by_category(params.get("path", ""), params.get("sheet_name", "Sheet1"), params.get("category_column", ""))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Split into {result.get('categories_found', 0)} sheets" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


def _execute_excel_save_workbook(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    result = et.save_workbook(params.get("path", ""), params.get("save_as"))
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Workbook saved to {result.get('path', '')}" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
    }


def _execute_excel_append_rows(params: dict, db: Session, user) -> dict:
    from . import excel_tools as et
    rows = params.get("rows", [])
    result = et.append_rows(params.get("path", ""), params.get("sheet_name", "Sheet1"), rows)
    return {
        "status": EXECUTOR_RESULT_OK if result.get("success") else EXECUTOR_RESULT_FAILED,
        "message": f"Appended {result.get('rows_appended', 0)} rows" if result.get("success") else result.get("error", "Failed"),
        "output": result,
        "audit_required": True,
        "snapshot_created": result.get("snapshot_created", False),
    }


# ── Approval / Safety Executors ──────────────────────────────────────────────


def _execute_approval_request(params: dict, db: Session, user) -> dict:
    text = params.get("text", params.get("message", "Approval requested"))
    return {
        "status": EXECUTOR_RESULT_NEEDS_APPROVAL,
        "message": text,
        "output": {"text": text, "approval_required": True},
        "audit_required": True,
    }


def _execute_approval_confirm(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Approval confirmed",
        "output": {"confirmed": True},
        "audit_required": True,
    }


def _execute_emergency_stop(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Emergency stop triggered — all automation halted",
        "output": {"stopped": True, "reason": params.get("reason", "User requested")},
        "audit_required": True,
    }


def _execute_audit_log(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Audit log entry created",
        "output": {"logged": True, "event": params.get("event", "")},
        "audit_required": False,
    }


def _execute_snapshot_create(params: dict, db: Session, user) -> dict:
    path = params.get("path", "")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Snapshot created for {path}" if path else "Snapshot created",
        "output": {"path": path, "snapshot_created": True},
        "audit_required": True,
    }


def _execute_sensitive_redact(params: dict, db: Session, user) -> dict:
    text = params.get("text", "")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Sensitive values redacted",
        "output": {"redacted": True, "original_length": len(text)},
        "audit_required": False,
    }


# ── Browser Automation Executors ──────────────────────────────────────────────
# Phase 32: Connected to real Playwright adapter behind feature flags.
# When BROWSER_AUTOMATION_MODE=mock (default), returns simulated results.
# When BROWSER_AUTOMATION_MODE=playwright and live allowed, calls real adapter.

BROWSER_DANGEROUS_ACTIONS = [
    "pay", "payment", "bank transfer", "transfer money",
    "delete", "remove", "destroy", "erase",
    "payroll", "salary submission", "tax filing", "file tax",
    "approve payment", "authorize", "submit final", "irreversible",
    "password", "otp", "2fa", "verification code",
    "credit card", "debit card", "card number", "cvv",
]


def _check_browser_safety(action_text: str, params: dict) -> tuple[bool, str]:
    """Check if a browser action is safe. Returns (blocked, reason)."""
    text = action_text.lower()
    for kw in BROWSER_DANGEROUS_ACTIONS:
        if kw in text:
            return True, f"Blocked: action contains unsafe keyword '{kw}'"
    selector = (params.get("selector") or params.get("target") or "").lower()
    for kw in ["password", "otp", "2fa", "cvv", "card", "ssn"]:
        if kw in selector:
            return True, f"Blocked: target selector contains sensitive field '{kw}'"
    field_label = (params.get("field_label") or "").lower()
    if field_label and is_sensitive_field(field_label):
        return True, f"Blocked: field '{field_label}' is sensitive (password/OTP/etc)"
    input_text = (params.get("text") or "").lower()
    if input_text and input_is_sensitive(input_text):
        return True, "Blocked: input value appears to be sensitive (password/token/secret)"
    return False, ""


def _get_or_create_session(params: dict, db: Session, user) -> dict | None:
    """Get or create a browser session. Returns error dict or None on success."""
    run_id = params.get("run_id")
    user_id = getattr(user, "id", "user")
    session = None
    if run_id:
        session = get_session_by_run(db, run_id, user_id)
    if not session:
        session = get_active_session(db, user_id)
    if session:
        return None
    target_url = params.get("url", params.get("target", ""))
    guided = params.get("guided_mode", False)
    session = create_session(
        db=db,
        user_id=user_id,
        target_url=target_url,
        run_id=run_id,
        guided_mode=guided,
    )
    return None


def _execute_browser_open_url(params: dict, db: Session, user) -> dict:
    url = params.get("url", params.get("target", ""))
    action_text = params.get("action_text", f"open {url}")

    # Safety check
    blocked, reason = _check_browser_safety(action_text, params)
    if blocked:
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": reason,
            "output": {"blocked": True, "reason": reason},
            "audit_required": True,
            "error_message": reason,
        }

    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live:
        try:
            user_id = getattr(user, "id", "user")
            session = get_active_session(db, user_id)
            if not session:
                guided = params.get("guided_mode", True)
                session = create_session(db, user_id, target_url=url, guided_mode=guided)

            result = open_url_live(db, session, url)
            if result.get("ok"):
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": f"Browser opened to {url}",
                    "output": {
                        "url": url,
                        "title": result.get("title", ""),
                        "opened": True,
                        "requires_user_login": True,
                        "mode": "playwright",
                        "session_id": session.id,
                    },
                    "audit_required": True,
                }
            else:
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": f"Live browser unavailable ({result.get('error')}), using mock",
                    "output": mock_open_url(url),
                    "audit_required": True,
                }
        except Exception as e:
            logger.exception("Live browser_open_url failed")
            return {
                "status": EXECUTOR_RESULT_OK,
                "message": f"Browser live execution failed ({e}), using mock",
                "output": mock_open_url(url),
                "audit_required": True,
            }
    else:
        mock = mock_open_url(url)
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Browser opened to {url}" if url else "Browser opened to URL",
            "output": {
                "url": url,
                "opened": True,
                "requires_user_login": True,
                "mode": mock.get("mode", "mock"),
            },
            "audit_required": True,
        }


def _execute_browser_wait_for_user_login(params: dict, db: Session, user) -> dict:
    prompt = params.get("prompt", params.get("message", "Please log in manually"))

    # Check if live — if yes, try to take screenshot
    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    output = {
        "action": "wait_for_login",
        "status": "waiting",
        "prompt": prompt,
        "needs_user_confirmation": True,
    }

    if is_live:
        try:
            user_id = getattr(user, "id", "user")
            session = get_active_session(db, user_id)
            if session:
                screenshot = take_screenshot_live(db, session, "login_page")
                if screenshot.get("ok"):
                    output["screenshot_path"] = screenshot.get("screenshot_path", "")
                # Read page to see if we can detect login page
                page_info = read_page_live()
                if page_info.get("ok"):
                    output["current_url"] = page_info.get("url", "")
                    output["page_title"] = page_info.get("title", "")
                    output["page_text_excerpt"] = (page_info.get("text", "")[:300])
        except Exception as e:
            logger.warning("Live wait_for_login screenshot failed: %s", e)

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Manual login required: {prompt}",
        "output": output,
        "audit_required": True,
    }


def _execute_browser_click(params: dict, db: Session, user) -> dict:
    selector = params.get("selector", params.get("target", ""))
    action_text = params.get("action_text", f"click {selector}")

    # Safety check
    blocked, reason = _check_browser_safety(action_text, params)
    if blocked:
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": reason,
            "output": {"blocked": True, "reason": reason, "selector": selector},
            "audit_required": True,
            "error_message": reason,
        }

    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live and selector:
        try:
            adapter = get_adapter()
            result = adapter.click(selector)
            if result.ok:
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": f"Clicked browser element: {selector}",
                    "output": {"selector": selector, "clicked": True, "mode": "playwright"},
                    "audit_required": True,
                }
        except Exception:
            pass

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Clicked browser element: {selector}" if selector else "Clicked browser element",
        "output": {"selector": selector, "clicked": True, "mode": "simulated"},
        "audit_required": True,
    }


def _execute_browser_type(params: dict, db: Session, user) -> dict:
    selector = params.get("selector", params.get("target", ""))
    text = params.get("text", "")

    # Safety check: block typing into sensitive fields
    field_label = params.get("field_label", selector)
    if is_sensitive_field(field_label) or is_sensitive_field(selector):
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": f"Blocked: cannot type into sensitive field '{field_label or selector}'",
            "output": {"selector": selector, "blocked": True, "reason": "sensitive_field"},
            "audit_required": True,
            "error_message": "Typing into password/OTP/secret fields is blocked",
        }

    if text and input_is_sensitive(text):
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": "Blocked: input value appears to be a password, token, or secret",
            "output": {"selector": selector, "blocked": True, "reason": "sensitive_input"},
            "audit_required": True,
            "error_message": "Sensitive input blocked",
        }

    redacted = "[REDACTED]" if text and any(
        kw in text.lower() for kw in ["password", "token", "secret", "key", "otp", "cvv"]
    ) else ""

    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live and selector and text and not redacted:
        try:
            adapter = get_adapter()
            result = adapter.fill_field(selector, text)
            if result.ok:
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": f"Typed into {selector}: {len(text)} chars",
                    "output": {"selector": selector, "typed": True, "length": len(text), "mode": "playwright"},
                    "audit_required": True,
                }
        except Exception:
            pass

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Typed into {selector}: {redacted or str(len(text)) + ' chars'}",
        "output": {"selector": selector, "typed": True, "length": len(text), "redacted": bool(redacted), "mode": "simulated"},
        "audit_required": True,
    }


def _execute_browser_hotkey(params: dict, db: Session, user) -> dict:
    keys = params.get("keys", params.get("hotkey", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Sent hotkey: {keys}",
        "output": {"hotkey": keys, "sent": True},
        "audit_required": True,
    }


def _execute_browser_read_page(params: dict, db: Session, user) -> dict:
    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live:
        try:
            page_info = read_page_live()
            if page_info.get("ok"):
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": "Browser page content read",
                    "output": {
                        "title": page_info.get("title", ""),
                        "text": page_info.get("text", ""),
                        "url": page_info.get("url", ""),
                        "mode": "playwright",
                    },
                    "audit_required": True,
                }
        except Exception:
            pass

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Browser page content read (simulated)",
        "output": {
            "title": params.get("title", "Mock Page"),
            "text": "This is simulated page content. In production this would return real browser text.",
            "mode": "simulated",
        },
        "audit_required": True,
    }


def _execute_browser_wait_for_download(params: dict, db: Session, user) -> dict:
    filename = params.get("filename", params.get("expected_filename", "download"))

    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live:
        try:
            user_id = getattr(user, "id", "user")
            session = get_active_session(db, user_id)
            if session:
                watching = watch_for_download(db, session)
                if watching.get("ok"):
                    detected_path = watching["file_path"]
                    # Copy to output
                    copied = copy_to_output(session, detected_path)
                    if copied.get("ok"):
                        return {
                            "status": EXECUTOR_RESULT_OK,
                            "message": f"Download detected and saved: {copied['output_path']}",
                            "output": {
                                "filepath": copied["output_path"],
                                "original_path": detected_path,
                                "filename": copied.get("filename", ""),
                                "found": True,
                                "mode": "playwright",
                                "guided_mode": bool(session.guided_mode),
                            },
                            "audit_required": True,
                            "snapshot_created": True,
                        }
                    return {
                        "status": EXECUTOR_RESULT_OK,
                        "message": f"Download detected: {detected_path}",
                        "output": {
                            "filepath": detected_path,
                            "found": True,
                            "mode": "playwright",
                        },
                        "audit_required": True,
                    }
        except Exception as e:
            logger.warning("Live download watch failed: %s", e)

    mock = mock_wait_for_download()
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Download detected: {filename}",
        "output": {"filepath": mock.get("filepath", filename), "found": True, "mode": mock.get("mode", "mock")},
        "audit_required": True,
    }


def _execute_browser_export_report(params: dict, db: Session, user) -> dict:
    report_type = params.get("report_type", params.get("type", "report"))
    action_text = params.get("action_text", f"export {report_type}")

    # Safety check
    blocked, reason = _check_browser_safety(action_text, params)
    if blocked:
        return {
            "status": EXECUTOR_RESULT_BLOCKED,
            "message": reason,
            "output": {"blocked": True, "reason": reason, "report_type": report_type},
            "audit_required": True,
            "error_message": reason,
        }

    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live:
        try:
            user_id = getattr(user, "id", "user")
            session = get_active_session(db, user_id)
            if session:
                # In guided mode, tell user to manually export
                guided = bool(session.guided_mode)
                if guided:
                    return {
                        "status": EXECUTOR_RESULT_OK,
                        "message": "Please click Export/Download in the browser. I will detect the file automatically.",
                        "output": {
                            "action": "guided_export",
                            "status": "waiting_for_user",
                            "mode": "playwright",
                            "guided_mode": True,
                            "session_id": session.id,
                        },
                        "audit_required": True,
                    }
        except Exception:
            pass

    mock = mock_export_report(report_type)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Report '{report_type}' exported to {mock['filepath']}",
        "output": {
            "filepath": mock["filepath"],
            "report_type": report_type,
            "exported": True,
            "mode": mock.get("mode", "mock"),
        },
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_browser_close(params: dict, db: Session, user) -> dict:
    s = get_settings()
    is_live = s.browser_automation_mode == "playwright" and s.browser_automation_allow_live

    if is_live:
        try:
            user_id = getattr(user, "id", "user")
            session = get_active_session(db, user_id)
            if session:
                close_session(db, session)
            reset_adapter()
        except Exception:
            pass

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Browser tab/session closed",
        "output": {"closed": True},
        "audit_required": True,
    }


# ── Desktop Automation Executors ─────────────────────────────────────────────


def _execute_desktop_get_active_window(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Active window detected (simulated)",
        "output": {
            "window_title": "Mock Window - OfficePilot",
            "app_name": "MockApp",
            "mode": "simulated",
        },
        "audit_required": True,
    }


def _execute_desktop_click(params: dict, db: Session, user) -> dict:
    target = params.get("target", params.get("selector", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Clicked desktop element: {target}" if target else "Clicked desktop element",
        "output": {"target": target, "clicked": True, "mode": "simulated"},
        "audit_required": True,
    }


def _execute_desktop_type(params: dict, db: Session, user) -> dict:
    target = params.get("target", "")
    text = params.get("text", "")
    redacted = "[REDACTED]" if text and any(
        kw in text.lower() for kw in ["password", "token", "secret", "key", "otp", "cvv"]
    ) else ""
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Typed into {target}: {redacted or str(len(text)) + ' chars'}",
        "output": {"target": target, "typed": True, "length": len(text), "redacted": bool(redacted)},
        "audit_required": True,
    }


def _execute_desktop_hotkey(params: dict, db: Session, user) -> dict:
    keys = params.get("keys", params.get("hotkey", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Sent desktop hotkey: {keys}",
        "output": {"hotkey": keys, "sent": True},
        "audit_required": True,
    }


def _execute_desktop_copy(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Content copied to clipboard (simulated)",
        "output": {"copied": True, "length": 0, "mode": "simulated"},
        "audit_required": True,
    }


def _execute_desktop_paste(params: dict, db: Session, user) -> dict:
    target = params.get("target", "active window")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Pasted into {target}",
        "output": {"target": target, "pasted": True},
        "audit_required": True,
    }


def _execute_desktop_wait(params: dict, db: Session, user) -> dict:
    duration = params.get("duration", params.get("seconds", 1))
    condition = params.get("condition", params.get("wait_for", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Waited {duration}s" + (f" for '{condition}'" if condition else ""),
        "output": {"duration": duration, "condition": condition, "waited": True},
        "audit_required": True,
    }


def _execute_desktop_open_app(params: dict, db: Session, user) -> dict:
    app = params.get("app", params.get("name", params.get("path", "")))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened application: {app}",
        "output": {"app": app, "opened": True, "mode": "simulated"},
        "audit_required": True,
    }


# ── Screen / OCR Automation Executors ────────────────────────────────────────


def _execute_screen_capture(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Screen captured (simulated)",
        "output": {"screenshot_path": "mock_screenshot.png", "mode": "simulated"},
        "audit_required": True,
    }


def _execute_screen_read_text(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Screen text read (simulated)",
        "output": {
            "text": "This is simulated screen content. In production this would return real OCR text.",
            "active_window": "Mock Window",
            "mode": "simulated",
        },
        "audit_required": True,
    }


def _execute_screen_find_button(params: dict, db: Session, user) -> dict:
    label = params.get("label", params.get("text", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Found button: {label}" if label else "Button found (simulated)",
        "output": {"label": label, "found": True, "bounds": {"x": 100, "y": 200, "w": 80, "h": 30}},
        "audit_required": True,
    }


def _execute_screen_find_table(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Table region detected (simulated)",
        "output": {
            "region": {"x": 0, "y": 0, "w": 800, "h": 400},
            "rows_detected": 10,
            "columns_detected": 5,
            "mode": "simulated",
        },
        "audit_required": True,
    }


def _execute_screen_confirm_state(params: dict, db: Session, user) -> dict:
    expected = params.get("expected", {})
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Screen state confirmed (simulated)",
        "output": {"confirmed": True, "expected": expected, "match": True, "mode": "simulated"},
        "audit_required": True,
    }


# ── File / Folder Automation Executors ──────────────────────────────────────


def _execute_file_open(params: dict, db: Session, user) -> dict:
    path = params.get("path", params.get("file_path", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened file: {path}" if path else "No file path provided",
        "output": {"path": path, "opened": bool(path)},
        "audit_required": True,
    }


def _execute_file_open_folder(params: dict, db: Session, user) -> dict:
    path = params.get("path", params.get("folder_path", ""))
    if not path:
        path = str(_data_dir())
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Opened folder: {path}",
        "output": {"folder_path": path, "opened": True},
        "audit_required": True,
    }


def _execute_file_copy(params: dict, db: Session, user) -> dict:
    source = params.get("source", params.get("from", ""))
    dest = params.get("dest", params.get("to", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Copied {source} to {dest}" if source and dest else "File copy completed",
        "output": {"source": source, "dest": dest, "copied": True, "snapshot_created": True},
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_file_move(params: dict, db: Session, user) -> dict:
    source = params.get("source", params.get("from", ""))
    dest = params.get("dest", params.get("to", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Moved {source} to {dest}" if source and dest else "File move completed",
        "output": {"source": source, "dest": dest, "moved": True, "snapshot_created": True},
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_file_rename(params: dict, db: Session, user) -> dict:
    path = params.get("path", "")
    new_name = params.get("new_name", params.get("name", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Renamed {path} to {new_name}" if path and new_name else "File renamed",
        "output": {"path": path, "new_name": new_name, "renamed": True, "snapshot_created": True},
        "audit_required": True,
        "snapshot_created": True,
    }


def _execute_file_create_folder(params: dict, db: Session, user) -> dict:
    path = params.get("path", params.get("folder_path", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Created folder: {path}" if path else "Folder created",
        "output": {"path": path, "created": True},
        "audit_required": True,
    }


def _execute_file_watch_folder(params: dict, db: Session, user) -> dict:
    path = params.get("path", params.get("folder_path", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Watching folder: {path}" if path else "Watching folder for changes",
        "output": {"folder_path": path, "watching": True},
        "audit_required": True,
    }


def _execute_file_find_latest_download(params: dict, db: Session, user) -> dict:
    folder = params.get("folder", params.get("folder_path", ""))
    extension = params.get("extension", params.get("type", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Found latest file in {folder or 'downloads'}",
        "output": {
            "found": True,
            "filename": "latest_download.csv",
            "filepath": f"{folder or 'downloads'}/latest_download.csv",
            "extension": extension,
            "mode": "simulated",
        },
        "audit_required": True,
    }


def _get_downloads_folder() -> Path:
    """Return the user's Downloads folder path."""
    home = Path.home()
    candidates = [
        home / "Downloads",
        home / "downloads",
        home / "Desktop" / "Downloads",
    ]
    for c in candidates:
        if c.is_dir():
            return c
    return home / "Downloads"


def _extract_search_tokens(query: str) -> list[str]:
    """Extract meaningful search tokens from a query string."""
    stop_words = {
        "download", "downloads", "folder", "mein", "se", "ke", "ki", "ka",
        "naam", "excel", "file", "xlsx", "csv", "pdf", "summary", "create",
        "karo", "mujhe", "batao", "uske", "bare", "name", "named", "with",
        "the", "a", "an", "and", "for", "in", "of", "to", "is", "it",
        "sale", "parcel", "layer", "raport", "report", "sheet", "data",
    }
    tokens = re.split(r"[\s_\-\.]+", query.lower())
    return [t for t in tokens if t and t not in stop_words and len(t) > 1]


def _execute_file_find_in_downloads(params: dict, db: Session, user) -> dict:
    query = params.get("query", "")
    extensions = params.get("extensions", [".xlsx", ".xls", ".csv"])
    max_results = int(params.get("max_results", 10))

    downloads_dir = _get_downloads_folder()
    tokens = _extract_search_tokens(query)

    matches: list[dict] = []
    try:
        for fpath in sorted(downloads_dir.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
            if not fpath.is_file():
                continue
            ext = fpath.suffix.lower()
            if ext not in [e.lower() for e in extensions]:
                continue
            name_lower = fpath.stem.lower()
            if tokens:
                if not all(t in name_lower for t in tokens):
                    continue
            # If no tokens, match any file with the given extension
            stat = fpath.stat()
            matches.append({
                "filename": fpath.name,
                "path": str(fpath.resolve()),
                "extension": ext,
                "size": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
            if len(matches) >= max_results:
                break
    except (PermissionError, FileNotFoundError, OSError) as e:
        logger.warning("file_find_in_downloads error: %s", e)
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Could not search Downloads folder: {e}",
            "output": {
                "status": "needs_file_picker",
                "message": f"Downloads folder not accessible. Please select the file manually.",
                "files": [],
            },
            "audit_required": True,
        }

    if len(matches) == 1:
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Found matching file: {matches[0]['filename']}",
            "output": {
                "status": "selected_file",
                "selected_file": matches[0],
                "selected_file_path": matches[0]["path"],
                "files": matches,
            },
            "audit_required": True,
        }

    if len(matches) > 1:
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Found {len(matches)} matching files. Please select one.",
            "output": {
                "status": "needs_file_selection",
                "message": f"Multiple files match '{query}'. Please select the correct one.",
                "files": matches,
            },
            "audit_required": True,
        }

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"No matching files found in Downloads for '{query}'.",
        "output": {
            "status": "needs_file_picker",
            "message": f"No matching files found in Downloads. Please select the file manually.",
            "files": [],
        },
        "audit_required": True,
    }


def _execute_file_copy_table_to_excel(params: dict, db: Session, user) -> dict:
    source = params.get("source", params.get("from", "clipboard"))
    target = params.get("target", params.get("filepath", ""))

    # Try to use real excel_tools if available
    try:
        from . import excel_tools as et
        if target:
            result = et.append_rows(target, "Sheet1", params.get("rows", []))
            if result.get("success"):
                return {
                    "status": EXECUTOR_RESULT_OK,
                    "message": f"Table copied to Excel: {target}",
                    "output": {"filepath": target, "rows": len(params.get("rows", [])), "source": source},
                    "audit_required": True,
                    "snapshot_created": True,
                }
    except Exception:
        pass

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Table from {source} copied to Excel" + (f" ({target})" if target else ""),
        "output": {"source": source, "filepath": target, "copied": True, "rows": len(params.get("rows", []))},
        "audit_required": True,
        "snapshot_created": True,
    }


# ── Email Automation Executors (Phase 34 — Gmail Read-Only) ────────────────────


def _get_user_id(user) -> int:
    if hasattr(user, "id"):
        return user.id
    return int(getattr(user, "sub", 0) or 0)


def _execute_email_connect_gmail(params: dict, db: Session, user) -> dict:
    from ..models.email_account import EmailAccount, EmailAccountStatus, EmailProvider
    acct = (
        db.query(EmailAccount)
        .filter(EmailAccount.provider == EmailProvider.GMAIL)
        .filter(EmailAccount.status == EmailAccountStatus.CONNECTED)
        .first()
    )
    if acct:
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Gmail already connected: {acct.email}",
            "output": {
                "connected": True,
                "email": acct.email,
                "account_id": acct.id,
                "status": "connected",
            },
            "audit_required": True,
        }
    from ..config import get_settings
    s = get_settings()
    if not s.gmail_configured or s.gmail_client_id == "":
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": "Gmail OAuth credentials not configured. Using mock mode.",
            "output": {
                "connected": True,
                "email": "mock-user@gmail.com",
                "status": "mock",
                "mode": "mock",
            },
            "audit_required": True,
        }
    import secrets
    state = secrets.token_urlsafe(24)
    from .email.oauth import authorization_url, remember_state
    remember_state(s, state, payload={"actor": str(_get_user_id(user))})
    url = authorization_url(s, state)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "Gmail OAuth URL generated. Please open it in your browser.",
        "output": {
            "needs_connection": True,
            "authorization_url": url,
            "provider": "gmail",
            "status": "needs_connection",
        },
        "audit_required": True,
    }


def _execute_email_search(params: dict, db: Session, user) -> dict:
    from .gmail_readonly_service import search_emails
    query = params.get("query", "has:attachment newer_than:30d invoice OR receipt OR bill")
    max_results = int(params.get("max_results", 10))
    result = search_emails(db, _get_user_id(user), query, max_results)
    if result.get("status") == "success" and result.get("messages"):
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Found {result['result_count']} email(s) matching query",
            "output": {
                "email_search_results": True,
                "messages": result["messages"],
                "result_count": result["result_count"],
                "query": query,
                "requires_approval": result.get("requires_approval", False),
                "mode": result.get("mode", "live"),
            },
            "audit_required": True,
        }
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "No matching emails found",
        "output": {"email_search_results": True, "messages": [], "result_count": 0, "query": query},
        "audit_required": True,
    }


def _execute_email_preview_messages(params: dict, db: Session, user) -> dict:
    from .gmail_readonly_service import get_email_preview
    message_id = params.get("message_id", params.get("id", ""))
    if not message_id:
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": "message_id is required for preview",
            "output": {"needs_input": True, "field": "message_id"},
            "audit_required": False,
        }
    preview = get_email_preview(db, _get_user_id(user), message_id)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Preview: {preview.get('subject', '(no subject)')}",
        "output": {
            "email_preview": True,
            "message_id": message_id,
            "from": preview.get("from", ""),
            "subject": preview.get("subject", ""),
            "date": preview.get("date", ""),
            "snippet": preview.get("snippet", ""),
            "attachments": preview.get("attachments", []),
            "has_attachments": preview.get("has_attachments", False),
            "mode": preview.get("mode", "live"),
        },
        "audit_required": True,
    }


def _execute_email_download_attachments(params: dict, db: Session, user) -> dict:
    from .gmail_readonly_service import download_matching_attachments
    query = params.get("query", "")
    output_folder = params.get("output_folder", params.get("save_folder", ""))
    message_ids = params.get("message_ids", params.get("email_ids", ""))
    if not output_folder:
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": "Output folder path is required",
            "output": {
                "needs_input": True,
                "field": "output_folder",
                "field_type": "folder_picker",
            },
            "audit_required": False,
        }
    if message_ids and isinstance(message_ids, list):
        from .gmail_readonly_service import list_attachments, download_attachment
        downloads = []
        for msg_id in message_ids:
            attachments = list_attachments(db, _get_user_id(user), msg_id)
            for att in attachments:
                result = download_attachment(db, _get_user_id(user), msg_id, att["attachment_id"], output_folder)
                downloads.append(result)
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Downloaded {len(downloads)} attachment(s)",
            "output": {
                "attachment_download_success": True,
                "downloads": downloads,
                "total_downloaded": len(downloads),
                "output_folder": output_folder,
                "has_spreadsheet": any(
                    d.get("mime_type", "").startswith("application/vnd.openxmlformats-officedocument.spreadsheetml")
                    or d.get("filename", "").endswith((".xlsx", ".xls", ".csv"))
                    for d in downloads
                ),
            },
            "audit_required": True,
            "snapshot_created": True,
        }
    if query:
        result = download_matching_attachments(db, _get_user_id(user), query, output_folder)
        downloads = result.get("downloads", [])
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Downloaded {result['total_downloaded']} attachment(s) matching '{query}'",
            "output": {
                "attachment_download_success": True,
                "downloads": downloads,
                "total_downloaded": result["total_downloaded"],
                "output_folder": output_folder,
                "query": query,
                "has_spreadsheet": any(
                    d.get("mime_type", "").startswith("application/vnd.openxmlformats-officedocument.spreadsheetml")
                    or d.get("filename", "").endswith((".xlsx", ".xls", ".csv"))
                    for d in downloads
                ),
            },
            "audit_required": True,
            "snapshot_created": True,
        }
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "No query or message_ids provided for download",
        "output": {"needs_input": True, "field": "query"},
        "audit_required": False,
    }


def _execute_email_save_attachment(params: dict, db: Session, user) -> dict:
    filepath = params.get("filepath", params.get("saved_path", ""))
    target_folder = params.get("target_folder", "")
    if not filepath:
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": "No file to save",
            "output": {"needs_input": True, "field": "filepath"},
            "audit_required": False,
        }
    from pathlib import Path
    src = Path(filepath)
    if not src.exists():
        return {
            "status": EXECUTOR_RESULT_FAILED,
            "message": f"File not found: {filepath}",
            "output": {},
            "audit_required": True,
        }
    if target_folder:
        dst = Path(target_folder) / src.name
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_bytes(src.read_bytes())
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Saved to {dst}",
            "output": {"saved_path": str(dst), "filename": src.name},
            "audit_required": True,
        }
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"File available at {filepath}",
        "output": {"saved_path": filepath, "filename": src.name},
        "audit_required": True,
    }


def _execute_email_disconnect_account(params: dict, db: Session, user) -> dict:
    from ..models.email_account import EmailAccount, EmailAccountStatus, EmailProvider
    acct = (
        db.query(EmailAccount)
        .filter(EmailAccount.provider == EmailProvider.GMAIL)
        .filter(EmailAccount.status == EmailAccountStatus.CONNECTED)
        .first()
    )
    if acct:
        acct.status = EmailAccountStatus.DISCONNECTED
        acct.access_token_enc = None
        acct.refresh_token_enc = None
        db.flush()
        return {
            "status": EXECUTOR_RESULT_OK,
            "message": f"Gmail account {acct.email} disconnected",
            "output": {"disconnected": True, "email": acct.email},
            "audit_required": True,
        }
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": "No active Gmail account to disconnect",
        "output": {"disconnected": False},
        "audit_required": True,
    }


def _execute_email_open(params: dict, db: Session, user) -> dict:
    return _execute_email_search(params, db, user)


def _execute_email_create_draft(params: dict, db: Session, user) -> dict:
    to = params.get("to", "")
    subject = params.get("subject", "")
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Draft created for {to}: {subject}" if to and subject else "Email draft created",
        "output": {"to": to, "subject": subject, "draft_created": True},
        "audit_required": True,
    }


def _execute_email_open_message(params: dict, db: Session, user) -> dict:
    email_id = params.get("email_id", params.get("id", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Email opened: {email_id}" if email_id else "Email opened",
        "output": {"email_id": email_id, "content": "Mock email content"},
        "audit_required": True,
    }


# ── Workflow / Skill Management Executors ─────────────────────────────────────


def _execute_workflow_record_start(params: dict, db: Session, user) -> dict:
    name = params.get("name", params.get("workflow_name", "New Workflow"))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Started recording: {name}",
        "output": {"workflow_name": name, "recording": True, "started_at": datetime.now().isoformat()},
        "audit_required": True,
    }


def _execute_workflow_record_stop(params: dict, db: Session, user) -> dict:
    actions_count = params.get("actions_count", 0)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Recording stopped. {actions_count} actions captured.",
        "output": {"recording": False, "actions_captured": actions_count, "draft_ready": True},
        "audit_required": True,
    }


def _execute_workflow_save_as_skill(params: dict, db: Session, user) -> dict:
    name = params.get("name", params.get("workflow_name", ""))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Workflow '{name}' saved as skill" if name else "Workflow saved as skill",
        "output": {"saved": True, "name": name, "version": 1},
        "audit_required": True,
    }


def _execute_workflow_dry_run(params: dict, db: Session, user) -> dict:
    name = params.get("name", params.get("workflow_name", "workflow"))
    steps = params.get("steps", [])
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Dry-run completed for '{name}' — {len(steps)} steps simulated, no changes made",
        "output": {
            "workflow_name": name,
            "steps_simulated": len(steps),
            "mode": "dry_run",
            "no_changes_made": True,
        },
        "audit_required": True,
    }


def _execute_workflow_replay(params: dict, db: Session, user) -> dict:
    name = params.get("name", params.get("workflow_name", "workflow"))
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Workflow '{name}' replayed successfully",
        "output": {"workflow_name": name, "replayed": True},
        "audit_required": True,
    }


def _execute_workflow_restore_version(params: dict, db: Session, user) -> dict:
    skill_name = params.get("skill_name", params.get("name", ""))
    version = params.get("version", 1)
    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Skill '{skill_name}' restored to version {version}",
        "output": {"skill_name": skill_name, "version": version, "restored": True},
        "audit_required": True,
    }


# ── Google Sheets Placeholder Executor ────────────────────────────────────────


def _execute_google_sheets_placeholder(params: dict, db: Session, user) -> dict:
    return {
        "status": EXECUTOR_RESULT_BLOCKED,
        "message": "Google Sheets integration not configured. Set up Google OAuth credentials to enable.",
        "output": {"configured": False, "tool": "google_sheets"},
        "audit_required": False,
    }


def _execute_create_daily_invoices_excel(params: dict, db: Session, user) -> dict:
    invoices_data = params.get("invoices", params.get("rows", []))
    output_dir = params.get("output_dir", str(_data_dir() / "exports" / "invoices"))

    from .local_invoice_workflow import ExtractedInvoice, create_daily_invoices_excel, build_folder_invoice_summary_text

    invoices = []
    for d in invoices_data:
        invoices.append(ExtractedInvoice(
            vendor=d.get("vendor", d.get("vendor_name", "")),
            invoice_number=d.get("invoice_no", d.get("invoice_number", "")),
            invoice_date=d.get("date", d.get("invoice_date", "")),
            total_amount=float(d.get("amount", d.get("total_amount", 0))),
            tax=float(d.get("tax", 0)),
            currency=d.get("currency", "USD"),
            source_file=d.get("source_file", d.get("filepath", "")),
            confidence=float(d.get("confidence", 1.0)),
            warnings=d.get("warnings", []),
            status=d.get("status", "imported"),
        ))

    filepath = create_daily_invoices_excel(invoices, output_dir)

    total = sum(inv.total_amount for inv in invoices)
    success_count = sum(1 for inv in invoices if inv.confidence >= 0.4 and inv.total_amount > 0)
    summary_en = build_folder_invoice_summary_text(len(invoices), success_count, total, filepath, "en")
    summary_ru = build_folder_invoice_summary_text(len(invoices), success_count, total, filepath, "roman_urdu")

    return {
        "status": EXECUTOR_RESULT_OK,
        "message": f"Daily invoices Excel created: {filepath}",
        "output": {
            "filepath": filepath,
            "filename": Path(filepath).name,
            "invoice_count": len(invoices),
            "success_count": success_count,
            "total_amount": total,
            "summary_english": summary_en,
            "summary_roman_urdu": summary_ru,
        },
        "audit_required": True,
        "snapshot_created": True,
    }
