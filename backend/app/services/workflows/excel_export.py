"""Phase 6 — Graph 3: approved_invoice_export_graph.

Wraps the existing Excel export + folder organiser as a LangGraph
``StateGraph`` with a mandatory human-approval checkpoint before
any file is written.

Nodes:

1. ``select_approved_invoices`` — query all invoices in
   ``APPROVED`` status.
2. ``validate_export_ready``    — sanity check (no missing
   required fields, no duplicates in the batch).
3. ``human_approval_checkpoint``— pause; the runner creates a
   :class:`WorkflowApproval` row and stops until the user POSTs
   ``/approve`` (or ``/reject``).
4. ``write_excel``              — call ``build_excel`` from
   :mod:`app.services.excel_export`.
5. ``organize_files``           — move the original files to
   the configured folder layout.
6. ``mark_exported``            — flip the invoices' status to
   ``EXPORTED``.
7. ``audit_log``                — write the workflow-level audit
   row.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import TypedDict

from langgraph.graph import END, START, StateGraph

from ..excel_export import build_excel, mark_exported
from ..organizer import organize

logger = logging.getLogger(__name__)


class ExcelExportState(TypedDict, total=False):
    # ---- input ----
    actor: str

    # ---- intermediate ----
    invoice_ids: list            # ids in scope for export
    invoice_count: int
    missing_fields: list         # per-invoice list of missing fields
    duplicate_in_batch: list     # duplicate invoice numbers within the batch
    excel_path: str
    organized_count: int
    error: str


EXCEL_EXPORT_NODES = [
    "select_approved_invoices",
    "validate_export_ready",
    "human_approval_checkpoint",
    "write_excel",
    "organize_files",
    "mark_exported",
    "audit_log",
]


REQUIRED_FOR_EXPORT = ("vendor_name", "invoice_number", "invoice_date", "total_amount")


# ---------------------------------------------------------------- node handlers


def _node_select_approved_invoices(state, runner):
    from ...models.invoice import Invoice, InvoiceStatus
    db = runner.db
    rows = (
        db.query(Invoice)
        .filter(Invoice.status == InvoiceStatus.APPROVED)
        .order_by(Invoice.id.asc())
        .all()
    )
    return {
        "invoice_ids": [r.id for r in rows],
        "invoice_count": len(rows),
    }


def _node_validate_export_ready(state, runner):
    from ...models.invoice import Invoice
    db = runner.db
    rows = db.query(Invoice).filter(Invoice.id.in_(state.get("invoice_ids") or [])).all()
    missing: list[dict] = []
    seen_numbers: dict[str, int] = {}
    dup: list[dict] = []
    for inv in rows:
        for fld in REQUIRED_FOR_EXPORT:
            if getattr(inv, fld) in (None, "", 0.0) and fld != "total_amount":
                missing.append({"invoice_id": inv.id, "missing": fld})
        if getattr(inv, "total_amount", None) is None:
            missing.append({"invoice_id": inv.id, "missing": "total_amount"})
        if inv.invoice_number:
            seen_numbers.setdefault(inv.invoice_number, 0)
            seen_numbers[inv.invoice_number] += 1
    for num, count in seen_numbers.items():
        if count > 1:
            dup.append({"invoice_number": num, "count": count})
    if missing:
        return {
            "missing_fields": missing,
            "duplicate_in_batch": dup,
            "error": f"export validation failed: {len(missing)} missing field(s)",
        }
    return {"missing_fields": [], "duplicate_in_batch": dup}


def _node_human_approval_checkpoint(state, runner):
    """Pause the run and ask the user to confirm. We don't auto-
    approve — that would defeat the whole point of the trust
    layer."""
    summary = {
        "invoice_count": state.get("invoice_count", 0),
        "missing_fields": state.get("missing_fields", []),
        "duplicate_in_batch": state.get("duplicate_in_batch", []),
    }
    after_preview = {
        "excel_path_will_be": "exports/invoices_<timestamp>.xlsx",
        "files_will_be_organized": state.get("invoice_count", 0),
        "invoices_will_be_marked": "EXPORTED",
    }
    runner.create_approval(
        "human_approval_checkpoint",
        (
            f"Export {state.get('invoice_count', 0)} approved invoice(s) to Excel? "
            "Review the missing-fields and duplicates lists below."
        ),
        before=summary,
        after=after_preview,
    )
    # Returning {} is fine — the runner halts because we created an
    # approval row.
    return {}


def _node_write_excel(state, runner):
    """Write the approved invoices to an .xlsx file.

    The existing :func:`app.services.excel_export.build_excel` writes
    *all* approved invoices in the DB. For the workflow we want to
    scope the export to the invoice_ids we just selected, so we
    inline a small writer that produces the same column layout
    (we share the :data:`COLUMNS` constant) but only the rows in
    scope.
    """
    from ...models.invoice import Invoice, InvoiceStatus
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from ...services.excel_export import COLUMNS

    db = runner.db
    settings = runner.settings
    rows = (
        db.query(Invoice)
        .filter(Invoice.id.in_(state.get("invoice_ids") or []))
        .order_by(Invoice.id.asc())
        .all()
    )
    approved_rows = [r for r in rows if r.status == InvoiceStatus.APPROVED]
    if not approved_rows:
        return {"error": "no approved invoices to export"}

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(settings.exports_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"approved_invoices_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Invoices"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for r_idx, inv in enumerate(approved_rows, start=2):
        ws.cell(row=r_idx, column=1, value=inv.id)
        ws.cell(row=r_idx, column=2, value=inv.vendor_name or "")
        ws.cell(row=r_idx, column=3, value=inv.invoice_number or "")
        ws.cell(row=r_idx, column=4, value=inv.invoice_date or "")
        ws.cell(row=r_idx, column=5, value=inv.due_date or "")
        ws.cell(row=r_idx, column=6, value=inv.currency or "")
        ws.cell(row=r_idx, column=7, value=inv.subtotal)
        ws.cell(row=r_idx, column=8, value=inv.tax)
        ws.cell(row=r_idx, column=9, value=inv.total_amount)
    wb.save(str(out_path))
    return {"excel_path": str(out_path)}


def _node_organize_files(state, runner):
    from ...models.invoice import Invoice
    from ...models.invoice_file import InvoiceFile
    db = runner.db
    settings = runner.settings
    storage = Path(settings.storage_root)
    rules = _effective_rules(db, settings)
    pattern = rules.get("pattern", "{year}/{vendor}/{invoice_number}_{date}.pdf")
    strategy = rules.get("conflict_strategy", "suffix")
    rows = (
        db.query(Invoice)
        .filter(Invoice.id.in_(state.get("invoice_ids") or []))
        .all()
    )
    moved = 0
    for inv in rows:
        file_row: InvoiceFile = inv.file
        if not file_row or not file_row.stored_path:
            continue
        try:
            result = organize(
                storage_root=storage,
                invoice=inv,
                source_path=file_row.stored_path,
                pattern=pattern,
                conflict_strategy=strategy,
            )
            file_row.current_path = str(result.target)
            db.add(file_row)
            moved += 1
        except Exception as exc:
            logger.warning("organize failed for invoice %s: %s", inv.id, exc)
    db.commit()
    return {"organized_count": moved}


def _node_mark_exported(state, runner):
    from ...models.invoice import Invoice, InvoiceStatus
    db = runner.db
    ids = state.get("invoice_ids") or []
    mark_exported(db)
    # Refresh the list of invoices now marked EXPORTED for the audit row.
    rows = db.query(Invoice).filter(Invoice.id.in_(ids)).all()
    return {}


def _node_audit_log(state, runner):
    from ...services.audit import log_action
    if state.get("error"):
        return {}
    log_action(
        runner.db,
        actor=state.get("actor", "user"),
        action="workflow.export",
        entity_type="invoice",
        entity_id=None,
        details=(
            f"Workflow approved_invoice_export: count={state.get('invoice_count')}, "
            f"excel={state.get('excel_path')}, organized={state.get('organized_count')}"
        ),
        extra={
            "workflow_run_id": runner.run.id,
            "excel_path": state.get("excel_path"),
            "invoice_count": state.get("invoice_count"),
            "organized_count": state.get("organized_count"),
        },
    )
    return {}


# ----------------------------------------------------------- helpers


def _effective_rules(db, settings) -> dict:
    try:
        from ..organizer import get_effective_rules
        return get_effective_rules(db, settings) or {}
    except Exception:
        return {}


NODES = {
    "select_approved_invoices": _node_select_approved_invoices,
    "validate_export_ready": _node_validate_export_ready,
    "human_approval_checkpoint": _node_human_approval_checkpoint,
    "write_excel": _node_write_excel,
    "organize_files": _node_organize_files,
    "mark_exported": _node_mark_exported,
    "audit_log": _node_audit_log,
}


def build_excel_export_graph():
    g = StateGraph(ExcelExportState)
    for name, fn in NODES.items():
        g.add_node(name, fn)
    g.add_edge(START, "select_approved_invoices")
    for a, b in zip(EXCEL_EXPORT_NODES, EXCEL_EXPORT_NODES[1:]):
        g.add_edge(a, b)
    g.add_edge(EXCEL_EXPORT_NODES[-1], END)
    return g.compile(), NODES
