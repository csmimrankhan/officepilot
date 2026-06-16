"""Excel export of approved invoices (Phases 1-3)."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..config import Settings
from ..models.invoice import Invoice, InvoiceStatus

logger = logging.getLogger(__name__)


COLUMNS = [
    ("Invoice ID", 10),
    ("Vendor", 30),
    ("Invoice Number", 20),
    ("Invoice Date", 14),
    ("Due Date", 14),
    ("Currency", 10),
    ("Subtotal", 14),
    ("Tax", 12),
    ("Total", 14),
    ("Confidence", 12),
    ("Approved At", 20),
    ("Warnings", 50),
    ("Source File", 35),
    ("File Hash", 18),
]


def build_excel(
    db: Session,
    settings: Settings,
    include_rejected: bool = False,
    include_duplicate: bool = False,
) -> Path:
    """Build an .xlsx of all APPROVED invoices; return the file path.

    By default we only include approved rows. Admins can opt in to include
    rejected or duplicate rows for traceability, but they're written to a
    separate sheet so the main report stays clean.
    """
    settings.exports_dir.mkdir(parents=True, exist_ok=True)
    # Phase 10: snapshot the most recent previous approved_invoices
    # file before generating a new one. Each export is timestamped
    # so collisions are rare, but a snapshot gives the user an
    # "Undo last export" path even if they later delete the .xlsx
    # by hand. Snapshot only the freshest previous file (not the
    # whole history) to keep the snapshots dir small.
    from .snapshots import create_snapshot
    from ..models.file_snapshot import FileSnapshot

    prior = sorted(
        settings.exports_dir.glob("approved_invoices_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for prior_file in prior[:1]:
        try:
            snap = create_snapshot(
                source=prior_file,
                snapshots_root=settings.snapshots_dir,
                file_type="excel_export",
            )
            if snap is not None:
                db.add(
                    FileSnapshot(
                        file_type="excel_export",
                        original_path=str(prior_file),
                        snapshot_path=str(snap.snapshot_path),
                        action_type="excel_export.pre_export",
                        created_by="user",
                        restore_status="active",
                        file_hash_before=snap.file_hash,
                        size_bytes=snap.size_bytes,
                        notes="Snapshot of previous approved-invoices export",
                    )
                )
                db.flush()
        except Exception as _exc:  # pragma: no cover — best-effort
            logger.warning("Failed to snapshot prior export: %s", _exc)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out = settings.exports_dir / f"approved_invoices_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Invoices"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    approved_rows = (
        db.query(Invoice)
        .filter(Invoice.status == InvoiceStatus.APPROVED)
        .order_by(Invoice.id.asc())
        .all()
    )

    for r_idx, inv in enumerate(approved_rows, start=2):
        ws.cell(row=r_idx, column=1, value=inv.id)
        ws.cell(row=r_idx, column=2, value=inv.vendor_name or "")
        ws.cell(row=r_idx, column=3, value=inv.invoice_number or "")
        ws.cell(row=r_idx, column=4, value=inv.invoice_date or "")
        ws.cell(row=r_idx, column=5, value=inv.due_date or "")
        ws.cell(row=r_idx, column=6, value=inv.currency or "")
        ws.cell(row=r_idx, column=7, value=inv.subtotal)
        ws.cell(row=r_idx, column=8, value=inv.tax)
        ws.cell(row=r_idx, column=9, value=inv.total_amount)
        ws.cell(row=r_idx, column=10, value=inv.confidence_score)
        ws.cell(row=r_idx, column=11, value=inv.updated_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r_idx, column=12, value="; ".join(inv.warnings_json or []))
        ws.cell(row=r_idx, column=13, value=inv.file.original_filename if inv.file else "")
        ws.cell(row=r_idx, column=14, value=(inv.file.file_hash[:16] + "…") if inv.file else "")

    # Number formats
    for col in (7, 8, 9):
        for cell in ws[get_column_letter(col)][1:]:
            cell.number_format = "#,##0.00"
    ws[get_column_letter(10)][0].number_format = "0.00"

    ws.freeze_panes = "A2"

    # Optional admin-only traceability sheets.
    if include_rejected:
        _write_status_sheet(wb, db, "Rejected Invoices", InvoiceStatus.REJECTED, header_font, header_fill)
    if include_duplicate:
        _write_status_sheet(wb, db, "Duplicates", InvoiceStatus.DUPLICATE, header_font, header_fill)

    wb.save(out)
    return out


def _write_status_sheet(wb, db, title, status, header_font, header_fill) -> None:
    if title in wb.sheetnames:
        return
    ws = wb.create_sheet(title=title)
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    rows = (
        db.query(Invoice)
        .filter(Invoice.status == status)
        .order_by(Invoice.id.asc())
        .all()
    )
    for r_idx, inv in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=inv.id)
        ws.cell(row=r_idx, column=2, value=inv.vendor_name or "")
        ws.cell(row=r_idx, column=3, value=inv.invoice_number or "")
        ws.cell(row=r_idx, column=4, value=inv.invoice_date or "")
        ws.cell(row=r_idx, column=5, value=inv.due_date or "")
        ws.cell(row=r_idx, column=6, value=inv.currency or "")
        ws.cell(row=r_idx, column=7, value=inv.subtotal)
        ws.cell(row=r_idx, column=8, value=inv.tax)
        ws.cell(row=r_idx, column=9, value=inv.total_amount)
        ws.cell(row=r_idx, column=10, value=inv.confidence_score)
        ws.cell(row=r_idx, column=11, value=inv.updated_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r_idx, column=12, value="; ".join(inv.warnings_json or []))
        ws.cell(row=r_idx, column=13, value=inv.file.original_filename if inv.file else "")
        ws.cell(row=r_idx, column=14, value=(inv.file.file_hash[:16] + "…") if inv.file else "")


def mark_exported(db: Session, *, include_rejected: bool = False, include_duplicate: bool = False) -> List[int]:
    """Flip APPROVED invoices (and optionally the audit-only rows) to
    EXPORTED. Returns the IDs that were updated so the caller can log
    them in the audit log."""
    updated: list[int] = []
    q = db.query(Invoice).filter(Invoice.status == InvoiceStatus.APPROVED)
    for inv in q.all():
        inv.status = InvoiceStatus.EXPORTED
        updated.append(inv.id)
    if include_rejected:
        # We do not flip rejected invoices; admin opt-in is for the
        # traceability sheet only.
        pass
    return updated
