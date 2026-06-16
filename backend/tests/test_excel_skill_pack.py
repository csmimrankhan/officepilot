from __future__ import annotations

import json
import os
import tempfile

import pytest
from fastapi.testclient import TestClient


@pytest.fixture()
def sample_excel():
    import openpyxl
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Date", "Vendor", "Category", "Amount"])
    ws.append(["2024-01-15", "ABC Corp", "Office Supplies", 250.00])
    ws.append(["2024-01-20", "XYZ Ltd", "Software", 1200.00])
    ws.append(["2024-02-10", "ABC Corp", "Office Supplies", 180.00])
    ws.append(["2024-02-15", "Quick Books", "Accounting", 500.00])
    ws.append(["2024-03-05", "XYZ Ltd", "Software", 1200.00])
    wb.save(tmp.name)
    wb.close()
    yield tmp.name
    try:
        os.unlink(tmp.name)
    except OSError:
        pass

    bak_dir = os.path.join(os.path.dirname(tmp.name), "backup_before_auth_merge")
    if os.path.isdir(bak_dir):
        import shutil
        shutil.rmtree(bak_dir, ignore_errors=True)


# ── Excel Tools Unit Tests ───────────────────────────────────────────


class TestExcelTools:
    def test_create_workbook(self, sample_excel):
        from app.services.excel_tools import create_workbook
        result = create_workbook(sample_excel, sheets=["Sheet1", "Sheet2"])
        assert result["success"] is True
        assert "Sheet1" in result["sheets"]

    def test_read_sheet(self, sample_excel):
        from app.services.excel_tools import read_sheet
        result = read_sheet(sample_excel, "Sales")
        assert result["success"] is True
        assert result["row_count"] == 6
        assert result["rows"][0] == ["Date", "Vendor", "Category", "Amount"]

    def test_detect_columns(self, sample_excel):
        from app.services.excel_tools import detect_columns
        result = detect_columns(sample_excel, "Sales")
        assert result["success"] is True
        columns = result["columns"]
        assert len(columns) == 4
        assert columns[0]["header"] == "Date"
        assert columns[3]["header"] == "Amount"

    def test_apply_total_formula(self, sample_excel):
        from app.services.excel_tools import apply_total_formula
        result = apply_total_formula(sample_excel, "Sales", "D")
        assert result["success"] is True
        assert result["backup_created"] is True

    def test_add_total_row(self, sample_excel):
        from app.services.excel_tools import add_total_row
        result = add_total_row(sample_excel, "Sales", "D")
        assert result["success"] is True
        assert result["backup_created"] is True
        assert result["total_row"] > 0

    def test_create_summary_sheet(self, sample_excel):
        from app.services.excel_tools import create_summary_sheet
        result = create_summary_sheet(sample_excel, "Sales", "Vendor", "Amount")
        assert result["success"] is True
        assert result["summary_sheet"] is not None
        assert result["backup_created"] is True

    def test_compare_workbooks(self, sample_excel):
        from app.services.excel_tools import compare_workbooks
        result = compare_workbooks(sample_excel, sample_excel)
        assert result["success"] is True
        assert result["total_differences"] == 0

    def test_format_header(self, sample_excel):
        from app.services.excel_tools import format_header
        result = format_header(sample_excel, "Sales")
        assert result["success"] is True

    def test_auto_size_columns(self, sample_excel):
        from app.services.excel_tools import auto_size_columns
        result = auto_size_columns(sample_excel, "Sales")
        assert result["success"] is True

    def test_freeze_top_row(self, sample_excel):
        from app.services.excel_tools import freeze_top_row
        result = freeze_top_row(sample_excel, "Sales")
        assert result["success"] is True

    def test_backup_created_before_write(self, sample_excel):
        from app.services.excel_tools import create_backup_and_snapshot
        result = create_backup_and_snapshot(sample_excel)
        assert result["backup_created"] is True
        assert os.path.isfile(result["backup_path"])


# ── Formula Compatibility Tests ──────────────────────────────────────


class TestFormulaCompat:
    def test_excel_365_uses_xlookup(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("find", ExcelCompatibilityMode.EXCEL_365)
        assert "XLOOKUP" in formula

    def test_excel_2016_avoids_xlookup(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("find", ExcelCompatibilityMode.EXCEL_2016)
        assert "VLOOKUP" in formula or "INDEX" in formula

    def test_excel_2010_avoids_xlookup(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("find", ExcelCompatibilityMode.EXCEL_2010)
        assert "XLOOKUP" not in formula
        assert "VLOOKUP" in formula or "INDEX" in formula

    def test_sum_formula(self):
        from app.services.excel_formula_compat import choose_formula
        formula = choose_formula("sum")
        assert "SUM" in formula or "SUBTOTAL" in formula

    def test_unique_not_available_in_older_excel(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("unique", ExcelCompatibilityMode.EXCEL_2016)
        assert formula is None

    def test_unique_available_in_365(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("unique", ExcelCompatibilityMode.EXCEL_365)
        assert "UNIQUE" in formula

    def test_filter_not_available_in_2016(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("filter", ExcelCompatibilityMode.EXCEL_2016)
        assert formula is None

    def test_filter_available_in_365(self):
        from app.services.excel_formula_compat import choose_formula, ExcelCompatibilityMode
        formula = choose_formula("filter", ExcelCompatibilityMode.EXCEL_365)
        assert "FILTER" in formula

    def test_dangerous_formula_blocked(self):
        from app.services.excel_formula_compat import validate_formula_safety
        safe, reason = validate_formula_safety("=WEBSERVICE(\"http://evil.com\")")
        assert safe is False
        assert "WEBSERVICE" in reason

    def test_safe_formula_allowed(self):
        from app.services.excel_formula_compat import validate_formula_safety
        safe, reason = validate_formula_safety("=SUM(A1:A10)")
        assert safe is True


# ── Google Sheets Placeholder Tests ──────────────────────────────────


class TestGoogleSheetsPlaceholder:
    def test_create_sheet_not_configured(self):
        from app.services.excel_tools import google_sheets_create_sheet
        result = google_sheets_create_sheet("test")
        assert result["success"] is False
        assert "not configured" in result["error"].lower()

    def test_read_sheet_not_configured(self):
        from app.services.excel_tools import google_sheets_read_sheet
        result = google_sheets_read_sheet("id", "A1:B2")
        assert result["success"] is False
        assert "not configured" in result["error"].lower()


# ── Default Excel Skills Tests ───────────────────────────────────────


class TestDefaultAutomationSkills:
    def test_excel_summary_skill_exists(self):
        from app.services.accounting_skills import AUTOMATION_SKILL_TEMPLATES
        names = [s["name"] for s in AUTOMATION_SKILL_TEMPLATES]
        assert "Create Excel Summary" in names

    def test_pivot_skill_exists(self):
        from app.services.accounting_skills import AUTOMATION_SKILL_TEMPLATES
        names = [s["name"] for s in AUTOMATION_SKILL_TEMPLATES]
        assert "Create Pivot Table" in names

    def test_all_excel_skills_have_steps(self):
        from app.services.accounting_skills import AUTOMATION_SKILL_TEMPLATES
        for s in AUTOMATION_SKILL_TEMPLATES:
            assert len(s["steps"]) > 0
            assert len(s["trigger_phrases"]) > 0


# ── Agent Command Detection Tests ────────────────────────────────────


class TestExcelCommandDetection:
    def test_create_excel_summary_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("create excel summary")
        assert result is not None
        assert result["task_type"] == "excel_summary"

    def test_create_pivot_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("create pivot table by vendor")
        assert result is not None
        assert result["task_type"] == "excel_pivot"

    def test_apply_formula_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("add formula to total column")
        assert result is not None
        assert result["task_type"] == "excel_formula"

    def test_clean_csv_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("clean this csv")
        assert result is not None
        assert result["task_type"] == "excel_clean"

    def test_compare_excel_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("compare two excel files")
        assert result is not None
        assert result["task_type"] == "excel_compare"

    def test_non_excel_command_not_detected(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("show me pending invoices")
        assert result is None

    def test_excel_command_clarification_needed(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("create excel summary")
        assert result["clarification_needed"] is True
        assert "Excel file" in (result.get("clarification_question") or "")

    def test_summary_plan_has_steps(self):
        from app.services.accountant_autopilot import _check_excel_command
        result = _check_excel_command("create excel summary")
        steps = result.get("steps", [])
        assert len(steps) == 1
        assert steps[0]["tool"] == "excel_create_summary_from_file"
