"""Phase 45A — Automated Bank Reconciliation via Semantic Memory & COM tests."""

from __future__ import annotations

import json
import os
import tempfile
import time

os.environ["ALLOW_OPEN_REGISTRATION"] = "true"
os.environ["AGENT_PROVIDER"] = "mock"
os.environ["DATABASE_URL"] = "sqlite:///./test_phase45a_rec.db"
os.environ["OFFICEPILOT_APP_VERSION"] = "1.0.0"
os.environ["BANK_FEED_MODE"] = "mock"

import pytest
from fastapi.testclient import TestClient

from app.db import Base, SessionLocal, init_db
from app.main import app
from app.services.agent_tool_executor import execute_tool
from app.services.bank_reconciliation import (
    BankFeedAdapter,
    BankTransaction,
    ReconciliationEngine,
    ReconciliationRecord,
    generate_reconciliation_excel,
)
from app.services.semantic_memory import SemanticMemory, get_semantic_memory, reset_semantic_memory
from app.services.tool_registry import get_tool

FAKE_USER = type("FakeUser", (), {"id": 1, "email": "test@test.com", "role": "admin"})()

SAMPLE_CSV = """date,description,amount,type
2026-01-15,Invoice payment from Acme Corp,5000.00,credit
2026-01-20,Office supplies - OfficeMart,1200.00,debit
2026-02-01,Consulting fee - TechSolutions,3000.00,credit
2026-02-10,Software subscription,450.00,debit
2026-03-05,Payment received GlobalTech,8200.00,credit"""


@pytest.fixture(autouse=True)
def _clean_db():
    init_db()
    db = SessionLocal()
    try:
        for table in reversed(Base.metadata.sorted_tables):
            db.execute(table.delete())
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()
    reset_semantic_memory()


@pytest.fixture
def db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
def seeded_semantic_memory():
    sm = get_semantic_memory()
    sm.index_invoice("inv-1", "Acme Corp monthly invoice payment $5000", {"vendor": "Acme Corp", "total_amount": 5000, "user_id": 1})
    sm.index_invoice("inv-2", "OfficeMart supplies order December", {"vendor": "OfficeMart", "total_amount": 1200, "user_id": 1})
    sm.index_invoice("inv-3", "TechSolutions consulting engagement", {"vendor": "TechSolutions", "total_amount": 3000, "user_id": 1})
    sm.index_invoice("inv-4", "GlobalTech payment received", {"vendor": "GlobalTech", "total_amount": 8200, "user_id": 1})
    yield sm
    reset_semantic_memory()


# ── BankFeedAdapter Tests ────────────────────────────────────────────────────


class TestBankFeedAdapter:
    def test_parse_csv_text(self):
        adapter = BankFeedAdapter()
        txns = adapter.parse_feed_text(SAMPLE_CSV, "feed.csv")
        assert len(txns) == 5
        assert txns[0].date == "2026-01-15"
        assert txns[0].description == "Invoice payment from Acme Corp"
        assert txns[0].amount == 5000.0
        assert txns[0].txn_type == "credit"

    def test_parse_csv_from_file(self):
        adapter = BankFeedAdapter()
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8") as f:
            f.write(SAMPLE_CSV)
            f.flush()
            fname = f.name
        try:
            txns = adapter.parse_feed(fname)
            assert len(txns) == 5
        finally:
            os.unlink(fname)

    def test_parse_json_text(self):
        adapter = BankFeedAdapter()
        data = json.dumps({
            "transactions": [
                {"date": "2026-01-15", "description": "Payment from Acme", "amount": 5000, "type": "credit"},
                {"date": "2026-01-20", "description": "OfficeMart supplies", "amount": 1200, "type": "debit"},
            ]
        })
        txns = adapter.parse_feed_text(data, "feed.json")
        assert len(txns) == 2
        assert txns[1].amount == 1200.0

    def test_parse_empty_returns_empty(self):
        adapter = BankFeedAdapter()
        assert adapter.parse_feed_text("", "feed.csv") == []
        assert adapter.parse_feed_text("date,description,amount,type\n", "feed.csv") == []

    def test_parse_amount_with_symbols(self):
        adapter = BankFeedAdapter()
        assert adapter._parse_amount("$1,234.56") == 1234.56
        assert adapter._parse_amount("invalid") == 0.0
        assert adapter._parse_amount(42) == 42.0

    def test_parse_csv_various_column_names(self):
        adapter = BankFeedAdapter()
        csv_data = "Date,Desc,Amount,Txn_Type\n2026-02-01,Test payment,100.00,credit\n"
        txns = adapter.parse_feed_text(csv_data, "feed.csv")
        assert len(txns) == 1


# ── ReconciliationEngine Tests ──────────────────────────────────────────────


class TestReconciliationEngine:
    def test_reconcile_matches_high_confidence(self, seeded_semantic_memory):
        engine = ReconciliationEngine(semantic_memory=seeded_semantic_memory)
        txns = [BankTransaction("2026-01-15", "Invoice payment from Acme Corp", 5000.0, "credit")]
        records = engine.reconcile(txns, extracted_invoices=[{"vendor": "Acme Corp", "total_amount": 5000.0, "id": "inv-1"}], user_id=1)
        assert len(records) == 1
        assert records[0].status in ("matched", "fuzzy_match")
        assert records[0].confidence >= 0.5

    def test_reconcile_with_extracted_invoices_fallback(self, seeded_semantic_memory):
        engine = ReconciliationEngine(semantic_memory=seeded_semantic_memory)
        txns = [BankTransaction("2026-03-10", "Payment SomeCorp", 2500.0, "credit")]
        extracted = [{"vendor": "SomeCorp", "total_amount": 2500.0, "id": "ext-1"}]
        records = engine.reconcile(txns, extracted_invoices=extracted, user_id=1)
        assert len(records) == 1

    def test_reconcile_unmatched(self, seeded_semantic_memory):
        engine = ReconciliationEngine(semantic_memory=seeded_semantic_memory)
        txns = [BankTransaction("2026-04-01", "Random unknown expense", 99.99, "debit")]
        records = engine.reconcile(txns, user_id=1)
        assert len(records) == 1
        assert records[0].status == "unmatched"
        assert records[0].confidence < 0.5

    def test_reconcile_multiple_transactions(self, seeded_semantic_memory):
        engine = ReconciliationEngine(semantic_memory=seeded_semantic_memory)
        txns = [
            BankTransaction("2026-01-15", "Invoice payment from Acme Corp", 5000.0, "credit"),
            BankTransaction("2026-01-20", "Office supplies - OfficeMart", 1200.0, "debit"),
            BankTransaction("2026-04-01", "Mystery charge", 50.0, "debit"),
        ]
        records = engine.reconcile(txns, extracted_invoices=[
            {"vendor": "Acme Corp", "total_amount": 5000.0, "id": "inv-1"},
            {"vendor": "OfficeMart", "total_amount": 1200.0, "id": "inv-2"},
        ], user_id=1)
        assert len(records) == 3
        statuses = [r.status for r in records]
        assert "matched" in statuses or "fuzzy_match" in statuses
        assert "unmatched" in statuses

    def test_exact_match_confidence_same_amount(self):
        engine = ReconciliationEngine()
        txn = BankTransaction("2026-01-15", "Acme Corp payment", 5000.0, "credit")
        inv = {"vendor": "Acme Corp", "total_amount": 5000.0}
        conf = engine._exact_match_confidence(txn, inv)
        assert conf == 1.0

    def test_exact_match_confidence_no_vendor(self):
        engine = ReconciliationEngine()
        txn = BankTransaction("2026-01-15", "some description", 100.0, "debit")
        inv = {"total_amount": 100.0}
        conf = engine._exact_match_confidence(txn, inv)
        assert conf == 0.0


# ── Excel Report Generation Tests ────────────────────────────────────────────


class TestExcelGeneration:
    def test_generate_reconciliation_excel_openpyxl_fallback(self):
        records = [
            ReconciliationRecord("2026-01-15", "Payment Acme", 5000.0, "credit", "Acme Corp", 5000.0, "inv-1", 0.95, "matched"),
            ReconciliationRecord("2026-01-20", "OfficeMart", 1200.0, "debit", "OfficeMart", 1200.0, "inv-2", 0.75, "fuzzy_match"),
            ReconciliationRecord("2026-04-01", "Unknown", 50.0, "debit", None, None, None, 0.1, "unmatched"),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            result = generate_reconciliation_excel(records, out)
            assert result["status"] == "ok"
            assert result["records"] == 3
            assert os.path.getsize(out) > 0
        finally:
            try:
                os.unlink(out)
            except PermissionError:
                pass

    def test_generate_openpyxl_with_conditional_formatting(self):
        import openpyxl
        records = [
            ReconciliationRecord("2026-01-15", "Payment Acme", 5000.0, "credit", "Acme Corp", 5000.0, "inv-1", 0.95, "matched"),
            ReconciliationRecord("2026-01-20", "Unknown", 50.0, "debit", None, None, None, 0.1, "unmatched"),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            generate_reconciliation_excel(records, out)
            wb = openpyxl.load_workbook(out)
            ws = wb.active
            assert ws.title == "Reconciliation"
            assert ws.cell(row=1, column=1).value == "Bank Date"
            assert ws.cell(row=2, column=1).value == "2026-01-15"
            assert ws.cell(row=2, column=8).value == "matched"
            assert ws.cell(row=3, column=8).value == "unmatched"
            wb.close()
        finally:
            try:
                os.unlink(out)
            except PermissionError:
                pass


# ── Tool Registry Tests ─────────────────────────────────────────────────────


class TestToolRegistry:
    def test_bank_tools_registered(self):
        from app.services.tool_registry import TOOL_REGISTRY
        names = {t.name for t in TOOL_REGISTRY}
        assert "bank_parse_feed" in names
        assert "bank_reconcile_and_report" in names

    def test_bank_parse_feed_risk_low(self):
        tool = get_tool("bank_parse_feed")
        assert tool is not None
        assert tool.risk_level == "low"
        assert tool.approval_required is False

    def test_bank_reconcile_and_report_risk_medium(self):
        tool = get_tool("bank_reconcile_and_report")
        assert tool is not None
        assert tool.risk_level == "medium"
        assert tool.approval_required is True


# ── Executor Tests ───────────────────────────────────────────────────────────


class TestExecutors:
    def test_bank_parse_feed_executor(self, db):
        result = execute_tool(
            "bank_parse_feed",
            {"content": "date,description,amount,type\n2026-01-15,Test payment,100.00,credit"},
            "live",
            db,
            FAKE_USER,
        )
        assert result["status"] in ("success", "ok")
        assert result["output"]["count"] == 1

    def test_bank_parse_feed_no_params(self, db):
        result = execute_tool("bank_parse_feed", {}, "live", db, FAKE_USER)
        assert result["status"] in ("failed", "error")

    def test_bank_reconcile_and_report_generates_excel(self, db, seeded_semantic_memory):
        txns = [{"date": "2026-01-15", "description": "Invoice payment from Acme Corp", "amount": 5000.0, "type": "credit"}]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            result = execute_tool(
                "bank_reconcile_and_report",
                {"transactions": txns, "output_path": out},
                "live",
                db,
                FAKE_USER,
            )
            assert result["status"] in ("success", "ok")
            assert result["output"]["summary"]["total"] == 1
            assert os.path.getsize(out) > 0
        finally:
            try:
                os.unlink(out)
            except PermissionError:
                pass

    def test_bank_reconcile_no_transactions(self, db):
        result = execute_tool("bank_reconcile_and_report", {}, "live", db, FAKE_USER)
        assert result["status"] in ("failed", "error")

    def test_bank_reconcile_dry_run(self, db):
        result = execute_tool(
            "bank_reconcile_and_report",
            {"transactions": [{"date": "2026-01-15", "description": "Test", "amount": 100, "type": "credit"}]},
            "dry_run",
            db,
            FAKE_USER,
        )
        assert result["status"] == "dry_run"


# ── End-to-End Integration Tests ─────────────────────────────────────────────


class TestIntegration:
    def test_parse_then_reconcile(self, db, seeded_semantic_memory):
        parse_result = execute_tool(
            "bank_parse_feed",
            {"content": SAMPLE_CSV},
            "live",
            db,
            FAKE_USER,
        )
        assert parse_result["status"] in ("success", "ok")
        txns = parse_result["output"]["transactions"]
        assert len(txns) == 5

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            rec_result = execute_tool(
                "bank_reconcile_and_report",
                {"transactions": txns, "output_path": out},
                "live",
                db,
                FAKE_USER,
            )
            assert rec_result["status"] in ("success", "ok")
            summary = rec_result["output"]["summary"]
            assert summary["total"] == 5
            assert os.path.getsize(out) > 0
        finally:
            try:
                os.unlink(out)
            except PermissionError:
                pass

    def test_reconcile_endpoint(self, client, seeded_semantic_memory):
        from app.services.auth import hash_password, create_access_token
        from app.db import SessionLocal
        db = SessionLocal()
        try:
            from app.models.user import User
            existing = db.query(User).filter(User.email == "rec_test@test.com").first()
            if not existing:
                user = User(email="rec_test@test.com", password_hash=hash_password("pass"), role="admin")
                db.add(user)
                db.commit()
                db.refresh(user)
            token = create_access_token(user.id, "rec_test@test.com", "admin")
        finally:
            db.close()

        txns = [{"date": "2026-01-15", "description": "Invoice payment from Acme Corp", "amount": 5000.0, "type": "credit"}]
        resp = client.post(
            "/api/agent/plan-task",
            json={"command": "reconcile bank transactions", "context": {}},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code in (200, 422)
