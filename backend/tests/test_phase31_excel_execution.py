"""Phase 31 — Real Excel Automation Execution tests."""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

SAMPLE_ROWS = [
    ["Date", "Vendor", "Category", "Amount"],
    ["2026-06-01", "Acme", "Office Supplies", 1200],
    ["2026-06-02", "Beta", "Logistics", 850],
    ["2026-06-03", "Acme", "Office Supplies", 300],
    ["2026-06-04", "Delta", "Software", 1500],
]


def _make_sample(path: str) -> str:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    for row in SAMPLE_ROWS:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


# ═════════════════════════════════════════════════════════════════════════════
# PART 1 — excel_tools direct tests
# ═════════════════════════════════════════════════════════════════════════════


class TestValidateFile:
    def test_1_validates_xlsx(self):
        from app.services.excel_tools import validate_file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            result = validate_file(path)
            assert result["valid"] is True
            assert result["extension"] == ".xlsx"
        finally:
            os.unlink(path)

    def test_2_rejects_unsupported_extension(self):
        from app.services.excel_tools import validate_file
        result = validate_file(r"C:\fake.txt")
        assert result["valid"] is False
        assert "unsupported" in result.get("error", "").lower()

    def test_3_rejects_nonexistent_file(self):
        from app.services.excel_tools import validate_file
        result = validate_file(r"C:\nonexistent\file.xlsx")
        assert result["valid"] is False
        assert "not found" in result.get("error", "").lower()


class TestCreateBackup:
    def test_4_creates_backup_copy(self):
        from app.services.excel_tools import create_backup_and_snapshot
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        import openpyxl
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()
        try:
            result = create_backup_and_snapshot(path)
            assert result["backup_created"] is True
            assert os.path.isfile(result["backup_path"])
            assert result["backup_path"] != path
        finally:
            for p in [path, result.get("backup_path", "")]:
                if p and os.path.isfile(p):
                    os.unlink(p)

    def test_5_backup_fails_for_nonexistent(self):
        from app.services.excel_tools import create_backup_and_snapshot
        result = create_backup_and_snapshot(r"C:\nonexistent.xlsx")
        assert result["backup_created"] is False


class TestColumnDetection:
    def test_6_detects_basic_columns(self):
        from app.services.excel_tools import detect_columns
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = detect_columns(path, "Sales")
            assert result["success"] is True
            assert len(result["columns"]) == 4
            headers = [c["header"] for c in result["columns"]]
            assert "Date" in headers
            assert "Vendor" in headers
            assert "Category" in headers
            assert "Amount" in headers
        finally:
            os.unlink(path)

    def test_7_detects_accounting_semantics(self):
        from app.services.excel_tools import detect_columns_with_semantics
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = detect_columns_with_semantics(path, "Sales")
            assert result["success"] is True
            cols = {c["header"]: c["semantic_type"] for c in result["columns"]}
            assert cols.get("Date") == "date"
            assert cols.get("Vendor") == "vendor"
            assert cols.get("Category") == "category"
            assert cols.get("Amount") == "amount"
        finally:
            os.unlink(path)

    def test_8_suggests_summary_columns(self):
        from app.services.excel_tools import suggest_summary_columns
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = suggest_summary_columns(path, "Sales")
            assert result["success"] is True
            assert result["suggested_group_by"] is not None
            assert result["suggested_value"] is not None
        finally:
            os.unlink(path)


class TestSummaryFromFile:
    def test_9_dry_run_does_not_modify_file(self):
        from app.services.excel_tools import create_summary_from_file, get_column_letter
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        original_mtime = os.path.getmtime(path)
        try:
            result = create_summary_from_file(path, {"mode": "dry_run"})
            assert result["mode"] == "dry_run"
            assert result["would_modify_file"] is False
            assert "planned_actions" in result
            assert len(result["planned_actions"]) > 0
            # File not modified
            assert os.path.getmtime(path) == original_mtime
        finally:
            os.unlink(path)

    def test_10_live_run_creates_summary(self):
        from app.services.excel_tools import create_summary_from_file
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = create_summary_from_file(path, {"mode": "live"})
            assert result["status"] == "success"
            assert result["backup_path"] is not None
            assert os.path.isfile(result["backup_path"])
            assert result["output_path"] is not None
            assert os.path.isfile(result["output_path"])
            assert result["rows_processed"] == 4
            assert result["total"] == 3850.0  # 1200 + 850 + 300 + 1500
        finally:
            os.unlink(path)
            for key in ("backup_path", "output_path"):
                p = result.get(key)
                if p and os.path.isfile(p):
                    os.unlink(p)

    def test_11_original_file_unchanged(self):
        from app.services.excel_tools import create_summary_from_file
        import openpyxl
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        original_content = openpyxl.load_workbook(path, read_only=True)
        orig_sheets = original_content.sheetnames
        original_content.close()
        try:
            result = create_summary_from_file(path, {"mode": "live"})
            assert result["status"] == "success"
            # Original still has original sheets
            check = openpyxl.load_workbook(path, read_only=True)
            assert check.sheetnames == orig_sheets
            check.close()
        finally:
            os.unlink(path)
            for key in ("backup_path", "output_path"):
                p = result.get(key)
                if p and os.path.isfile(p):
                    os.unlink(p)

    def test_12_output_has_summary_with_correct_totals(self):
        from app.services.excel_tools import create_summary_from_file
        import openpyxl
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = create_summary_from_file(path, {"mode": "live"})
            assert result["status"] == "success"

            wb = openpyxl.load_workbook(result["output_path"])
            sheet_name = result["summary_sheet"]
            assert sheet_name in wb.sheetnames
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            # Header: [Category, Count, Total Amount, Average Amount]
            assert len(rows) >= 4  # header + 3 categories
            data = {str(r[0]): r[2] for r in rows[1:] if r[0] is not None and r[0] != "Grand Total"}
            assert data.get("Office Supplies") == 1500.0
            assert data.get("Logistics") == 850.0
            assert data.get("Software") == 1500.0
        finally:
            os.unlink(path)
            for key in ("backup_path", "output_path"):
                p = result.get(key)
                if p and os.path.isfile(p):
                    os.unlink(p)

    def test_13_failed_for_unsupported_file(self):
        from app.services.excel_tools import create_summary_from_file
        result = create_summary_from_file(r"C:\fake.txt", {"mode": "live"})
        assert result["status"] == "failed"


# ═════════════════════════════════════════════════════════════════════════════
# PART 2 — Executor integration tests
# ═════════════════════════════════════════════════════════════════════════════


class TestExecutorIntegration:
    def test_14_executor_returns_needs_input_when_file_missing(self):
        from app.services.agent_tool_executor import execute_tool
        from unittest.mock import MagicMock
        result = execute_tool("excel_create_summary_from_file", {}, "live", MagicMock(), MagicMock())
        assert result["status"] == "needs_approval"
        output = result.get("output", {})
        assert output.get("needs_input") is True
        assert output.get("input_type") == "file_picker"

    def test_15_executor_runs_summary_when_file_exists(self):
        from app.services.agent_tool_executor import execute_tool
        from unittest.mock import MagicMock
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            result = execute_tool("excel_create_summary_from_file", {
                "path": path,
                "mode": "live",
            }, "live", MagicMock(), MagicMock())
            assert result["status"] == "success"
            output = result.get("output", {})
            assert output.get("status") == "success"
            assert os.path.isfile(output["backup_path"])
            assert os.path.isfile(output["output_path"])
        finally:
            os.unlink(path)
            if result.get("output"):
                for key in ("backup_path", "output_path"):
                    p = result["output"].get(key)
                    if p and os.path.isfile(p):
                        os.unlink(p)

    def test_16_executor_dry_run_does_not_write(self):
        from app.services.agent_tool_executor import execute_tool
        from unittest.mock import MagicMock
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        original_mtime = os.path.getmtime(path)
        try:
            result = execute_tool("excel_create_summary_from_file", {
                "path": path,
                "mode": "dry_run",
            }, "dry_run", MagicMock(), MagicMock())
            assert result["status"] == "dry_run"
            assert os.path.getmtime(path) == original_mtime
        finally:
            os.unlink(path)

    def test_17_dangerous_formula_rejected(self):
        from app.services.excel_tools import apply_formula
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.append(["A"])
        wb.save(path)
        wb.close()
        try:
            result = apply_formula(path, "Sheet1", "B1", "=SHELL(\"rm -rf /\")")
            assert result.get("success") is False
            assert "dangerous" in str(result.get("error", "")).lower() or "unsafe" in str(result.get("error", "")).lower()
        finally:
            os.unlink(path)


# ═════════════════════════════════════════════════════════════════════════════
# PART 3 — Skill matching still works with updated skill
# ═════════════════════════════════════════════════════════════════════════════


class TestSkillMatching:
    def _register_and_login(self, client):
        uid = str(os.urandom(4).hex())
        resp = client.post("/api/auth/register", json={
            "email": f"excel31_{uid}@example.com",
            "password": "Test@1234",
            "full_name": "Excel31 Test",
        })
        token = resp.json()["access_token"]
        client.headers.update({"Authorization": f"Bearer {token}"})
        return client

    def test_18_create_excel_summary_still_matches(self, client):
        self._register_and_login(client)
        resp = client.post("/api/agent/plan-task", json={
            "command": "create excel summary",
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data.get("type") == "skill_match", f"Expected skill_match, got: {data.get('type')}"
        matched = data.get("matched_skill", {})
        assert matched.get("name") == "Create Excel Summary"
        steps = matched.get("steps", [])
        assert len(steps) > 0
        step_tools = [s["tool"] for s in steps]
        assert "excel_create_summary_from_file" in step_tools
        assert "read_pdf" not in step_tools

    def test_19_force_new_plan_still_works(self, client):
        self._register_and_login(client)
        resp = client.post("/api/agent/plan-task", json={
            "command": "create excel summary",
            "force_new_plan": True,
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data.get("type") != "skill_match"
        assert data.get("plan_id") is not None

    def test_20_force_new_plan_uses_consolidated_tool(self, client):
        self._register_and_login(client)
        resp = client.post("/api/agent/plan-task", json={
            "command": "create excel summary",
            "force_new_plan": True,
        })
        assert resp.status_code == 200
        steps = resp.json()["plan"].get("steps", [])
        assert len(steps) == 1
        assert steps[0]["tool"] == "excel_create_summary_from_file"

    def test_21_plan_flow_needs_input_when_no_file(self, client):
        self._register_and_login(client)
        r = client.post("/api/agent/plan-task", json={
            "command": "create excel summary",
            "force_new_plan": True,
        })
        plan_id = r.json()["plan_id"]
        r = client.post(f"/api/agent/plans/{plan_id}/approve", json={"mode": "live"})
        run_id = r.json()["run_id"]
        step_log_id = r.json()["steps"][0]["step_log_id"]
        r = client.post(f"/api/agent/runs/{run_id}/execute-step", json={
            "step_log_id": step_log_id,
        })
        assert r.status_code == 200
        result = r.json()
        assert result["step_status"] == "pending"
        output = result.get("result", {}).get("output", {})
        assert output.get("needs_input") is True
        assert output.get("input_type") == "file_picker"

    def test_22_plan_flow_executes_with_file(self, client):
        import tempfile, os
        from tests.test_phase31_excel_execution import _make_sample
        self._register_and_login(client)
        path = _make_sample(tempfile.mktemp(suffix=".xlsx"))
        try:
            r = client.post("/api/agent/plan-task", json={
                "command": "create excel summary",
                "force_new_plan": True,
            })
            plan_id = r.json()["plan_id"]
            r = client.post(f"/api/agent/plans/{plan_id}/approve", json={"mode": "live"})
            run_id = r.json()["run_id"]
            step_log_id = r.json()["steps"][0]["step_log_id"]
            r = client.post(f"/api/agent/runs/{run_id}/execute-step", json={
                "step_log_id": step_log_id,
                "file_path": path,
            })
            assert r.status_code == 200
            result = r.json()
            assert result["step_status"] == "completed"
            output = result.get("result", {}).get("output", {})
            assert output.get("status") == "success"
            assert output.get("total", 0) > 0
        finally:
            if os.path.isfile(path):
                os.unlink(path)
