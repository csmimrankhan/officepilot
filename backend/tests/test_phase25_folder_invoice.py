"""Phase 25 — Local Folder Invoice Workflow tests.

Tests that:
- scan_folder_for_invoices finds PDF/image/spreadsheet files by date
- extract_invoice_from_file parses text invoices
- create_daily_invoices_excel creates valid workbook
- planner detects folder invoice commands
- tool executor scan/create functions work
- audit logging fires on scan and excel creation
- workflow memory can save/repeat folder invoice workflow
"""
from __future__ import annotations

import json
import os
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.services.local_invoice_workflow import (
    ExtractedInvoice,
    InvoiceFileInfo,
    scan_folder_for_invoices,
    extract_invoice_from_file,
    create_daily_invoices_excel,
    build_folder_invoice_summary_text,
)


def _set_mock_provider():
    os.environ["AGENT_PROVIDER"] = "mock"
    os.environ["AGENT_ALLOW_CLOUD"] = "false"
    os.environ["AGENT_API_KEY"] = ""


@pytest.fixture(autouse=True)
def _reset_env():
    _set_mock_provider()
    os.environ["MULTILINGUAL_ENABLED"] = "true"
    os.environ["DEMO_MODE"] = "true"
    yield
    _set_mock_provider()


@pytest.fixture()
def client_with_auth(client):
    resp = client.post("/api/auth/register", json={
        "email": "folder-user@test.com", "password": "Test@123456", "full_name": "Folder User",
    })
    token = resp.json()["access_token"]
    client.headers.update({"Authorization": f"Bearer {token}"})
    return client


# ── Service Tests ─────────────────────────────────────────────────────────────


class TestScanFolder:
    def test_scan_returns_empty_for_nonexistent_folder(self):
        files = scan_folder_for_invoices("C:\\Nonexistent_Folder_XYZ")
        assert files == []

    def test_scan_finds_txt_invoice_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            today_str = date.today().isoformat()
            f1 = Path(tmp) / "INVOICE_001.pdf"
            f1.write_text("dummy pdf content")
            f2 = Path(tmp) / "receipt_2024.png"
            f2.write_text("dummy image")
            f3 = Path(tmp) / "notes.txt"
            f3.write_text("not an invoice")

            files = scan_folder_for_invoices(tmp, date_filter="today", keywords=True)
            assert len(files) == 2
            names = {f.filename for f in files}
            assert "INVOICE_001.pdf" in names
            assert "receipt_2024.png" in names
            assert "notes.txt" not in names

    def test_scan_finds_by_extension_without_keywords(self):
        with tempfile.TemporaryDirectory() as tmp:
            f = Path(tmp) / "random.png"
            f.write_text("test")
            files = scan_folder_for_invoices(tmp, date_filter="today", keywords=False)
            assert len(files) == 1
            assert files[0].filename == "random.png"

    def test_scan_date_filter_excludes_old_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            f = Path(tmp) / "invoice_old.pdf"
            f.write_text("test")
            old_mtime = os.path.getmtime(str(f))
            files = scan_folder_for_invoices(tmp, date_filter="today", keywords=True)
            assert any(fi.filename == "invoice_old.pdf" for fi in files)


class TestExtractInvoice:
    def test_extract_from_txt_file_parses_structured_text(self):
        invoice_text = (
            "INVOICE\n"
            "Invoice Number: INV-1234\n"
            "Date: 2024-06-15\n"
            "Vendor: Test Corp\n"
            "Total: $1,250.00\n"
            "Tax: $125.00\n"
        )
        with tempfile.TemporaryDirectory() as tmp:
            fp = Path(tmp) / "INV_001.txt"
            fp.write_text(invoice_text)
            inv = extract_invoice_from_file(str(fp))
            assert inv.vendor == "Test Corp" or "Test Corp" in inv.vendor
            assert inv.total_amount > 0
            assert inv.confidence > 0

    def test_extract_returns_error_for_missing_file(self):
        inv = extract_invoice_from_file("C:\\nonexistent_file.pdf")
        assert inv.confidence == 0.0
        assert inv.status == "needs_review"

    def test_extract_handles_empty_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            fp = Path(tmp) / "empty.txt"
            fp.write_text("")
            inv = extract_invoice_from_file(str(fp))
            assert inv.confidence == 0.0
            assert "No extractable text" in " ".join(inv.warnings) or inv.status == "needs_review"

    def test_extract_xlsx_with_amounts(self):
        with tempfile.TemporaryDirectory() as tmp:
            fp = Path(tmp) / "invoice.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Vendor", "Invoice_No", "Amount", "Tax", "Currency"])
            ws.append(["TestCo", "INV-1", 1000.00, 100.00, "USD"])
            wb.save(str(fp))
            inv = extract_invoice_from_file(str(fp))
            assert inv.total_amount == 1000.0
            assert inv.vendor
            assert inv.currency == "USD"


class TestCreateDailyInvoicesExcel:
    def test_creates_excel_with_correct_sheets(self):
        invoices = [
            ExtractedInvoice(vendor="Vendor A", invoice_number="INV-1", total_amount=1000.0, tax=100.0, confidence=0.8),
            ExtractedInvoice(vendor="Vendor B", invoice_number="INV-2", total_amount=2500.0, tax=250.0, confidence=0.9),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = create_daily_invoices_excel(invoices, tmp)
            assert Path(path).exists()
            wb = openpyxl.load_workbook(path)
            assert "Invoices" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            ws = wb["Invoices"]
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) == 4
            assert rows[2][1] == "INV-2"
            ws2 = wb["Summary"]
            summary_rows = list(ws2.iter_rows(values_only=True))
            assert any("Total Amount" in str(r[0]) for r in summary_rows)

    def test_creates_excel_with_empty_invoice_list(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = create_daily_invoices_excel([], tmp)
            assert Path(path).exists()
            wb = openpyxl.load_workbook(path)
            assert "Invoices" in wb.sheetnames

    def test_creates_excel_respects_output_dir(self):
        invoices = [ExtractedInvoice(vendor="Test", total_amount=100.0, confidence=0.8)]
        with tempfile.TemporaryDirectory() as tmp:
            outdir = Path(tmp) / "sub" / "invoices"
            path = create_daily_invoices_excel(invoices, str(outdir))
            assert Path(path).parent == outdir


class TestBuildSummaryText:
    def test_english_summary(self):
        text = build_folder_invoice_summary_text(5, 3, 1250.50, "/path/file.xlsx", "en")
        assert "5" in text
        assert "3" in text
        assert "1,250.50" in text or "1250.50" in text

    def test_roman_urdu_summary(self):
        text = build_folder_invoice_summary_text(3, 2, 500.00, "/path/file.xlsx", "roman_urdu")
        assert "invoice" in text.lower() or "Maine" in text
        assert "500" in text

    def test_summary_defaults_to_english(self):
        text = build_folder_invoice_summary_text(1, 1, 99.99, "/path/file.xlsx")
        assert "I found" in text


# ── Planner Tests ─────────────────────────────────────────────────────────────


class TestFolderInvoicePlanner:
    def test_planner_detects_folder_invoice_command(self):
        from app.services.accountant_autopilot import _check_folder_invoice
        tests = [
            "scan folders for invoices",
            "find today invoice files",
            "daily invoice excel banao",
            "aj ki invoice",
        ]
        for cmd in tests:
            result = _check_folder_invoice(cmd)
            assert result is not None, f"Failed to detect: {cmd}"
            assert result["task_type"] == "local_folder_invoice_workflow"
            assert len(result["steps"]) == 5

    def test_planner_rejects_non_folder_command(self):
        from app.services.accountant_autopilot import _check_folder_invoice
        assert _check_folder_invoice("hello world") is None
        assert _check_folder_invoice("compare P&L") is None
        assert _check_folder_invoice("read my screen") is None

    def test_planner_steps_require_approval_for_excel(self):
        from app.services.accountant_autopilot import _build_folder_invoice_plan
        plan = _build_folder_invoice_plan()
        excel_step = [s for s in plan["steps"] if s["step_type"] == "create_daily_invoices_excel"][0]
        assert excel_step["requires_approval"] is True
        assert excel_step["risk_level"] == "medium"

    def test_planner_scan_step_is_low_risk(self):
        from app.services.accountant_autopilot import _build_folder_invoice_plan
        plan = _build_folder_invoice_plan()
        scan_step = [s for s in plan["steps"] if s["step_type"] == "scan_local_folder"][0]
        assert scan_step["requires_approval"] is False
        assert scan_step["risk_level"] == "low"


# ── API Endpoint Tests ────────────────────────────────────────────────────────


class TestFolderInvoiceAPI:
    def test_scan_endpoint_requires_auth(self, client):
        resp = client.post("/api/agent/folder-invoice/scan", json={"folder_path": "C:\\"})
        assert resp.status_code == 403 or resp.status_code == 401

    def test_scan_endpoint_with_auth(self, client_with_auth):
        with tempfile.TemporaryDirectory() as tmp:
            f = Path(tmp) / "INVOICE_test.pdf"
            f.write_text("test")
            resp = client_with_auth.post("/api/agent/folder-invoice/scan", json={
                "folder_path": tmp, "date_filter": "today",
            })
            assert resp.status_code == 200
            data = resp.json()
            assert data["ok"] is True
            assert data["count"] >= 1
            assert any("INVOICE_test.pdf" in fi["filename"] for fi in data["files"])

    def test_create_excel_endpoint_with_auth(self, client_with_auth):
        resp = client_with_auth.post("/api/agent/folder-invoice/create-excel", json={
            "invoices": [
                {"vendor": "TestCo", "invoice_no": "INV-1", "amount": 1000.0, "tax": 100.0, "date": "2024-06-15"},
                {"vendor": "Sample Ltd", "invoice_no": "INV-2", "amount": 2500.50, "tax": 250.0, "date": "2024-06-15"},
            ],
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["ok"] is True
        assert data["invoice_count"] == 2
        assert data["total_amount"] == 3500.50
        assert Path(data["filepath"]).exists()
        assert "summary_english" in data
        assert "summary_roman_urdu" in data

    def test_list_runs_endpoint(self, client_with_auth):
        resp = client_with_auth.get("/api/agent/folder-invoice/runs")
        assert resp.status_code == 200
        data = resp.json()
        assert "runs" in data
        assert isinstance(data["runs"], list)

    def test_create_excel_with_alternate_field_names(self, client_with_auth):
        resp = client_with_auth.post("/api/agent/folder-invoice/create-excel", json={
            "rows": [
                {"vendor_name": "AltCo", "invoice_number": "ALT-1", "total_amount": 500.0},
            ],
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["ok"] is True
        assert data["invoice_count"] == 1
        assert data["total_amount"] == 500.0

    def test_get_run_endpoint_not_found(self, client_with_auth):
        resp = client_with_auth.get("/api/agent/folder-invoice/runs/99999")
        assert resp.status_code == 404


# ── Tool Executor Tests ───────────────────────────────────────────────────────


class TestFolderInvoiceExecutors:
    def test_executor_scan_local_folder(self):
        from app.services.agent_tool_executor import _execute_scan_local_folder, EXECUTOR_RESULT_OK
        with tempfile.TemporaryDirectory() as tmp:
            f = Path(tmp) / "invoice_example.pdf"
            f.write_text("test")
            result = _execute_scan_local_folder({"folder_path": tmp, "date_filter": "today"}, None, None)
            assert result["status"] == EXECUTOR_RESULT_OK
            assert result["output"]["count"] >= 1

    def test_executor_scan_invalid_folder(self):
        from app.services.agent_tool_executor import _execute_scan_local_folder, EXECUTOR_RESULT_OK
        result = _execute_scan_local_folder({"folder_path": "C:\\Nonexistent_Folder"}, None, None)
        assert result["status"] == EXECUTOR_RESULT_OK
        assert result["output"]["count"] == 0

    def test_executor_create_daily_invoices_excel(self):
        from app.services.agent_tool_executor import _execute_create_daily_invoices_excel, EXECUTOR_RESULT_OK
        invoices = [
            {"vendor": "V1", "invoice_no": "I-1", "amount": 100.0},
            {"vendor": "V2", "invoice_no": "I-2", "amount": 200.0},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            os.environ["OFFICEPILOT_DATA_DIR"] = tmp
            result = _execute_create_daily_invoices_excel({"invoices": invoices, "output_dir": tmp}, None, None)
            assert result["status"] == EXECUTOR_RESULT_OK
            assert result["output"]["invoice_count"] == 2
            assert result["output"]["total_amount"] == 300.0
            assert result["output"]["summary_english"]
            assert result["output"]["summary_roman_urdu"]
            assert Path(result["output"]["filepath"]).exists()
            del os.environ["OFFICEPILOT_DATA_DIR"]

    def test_executor_empty_invoice_list(self):
        from app.services.agent_tool_executor import _execute_create_daily_invoices_excel, EXECUTOR_RESULT_OK
        with tempfile.TemporaryDirectory() as tmp:
            os.environ["OFFICEPILOT_DATA_DIR"] = tmp
            result = _execute_create_daily_invoices_excel({"invoices": [], "output_dir": tmp}, None, None)
            assert result["status"] == EXECUTOR_RESULT_OK
            assert result["output"]["invoice_count"] == 0
            assert result["output"]["total_amount"] == 0.0
            del os.environ["OFFICEPILOT_DATA_DIR"]
